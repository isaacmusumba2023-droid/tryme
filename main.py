import streamlit as st
from supabase import create_client, Client
import pandas as pd
from datetime import datetime, date
import numpy as np
import base64
import os
import time

min_date = date(1900, 1, 1)
max_date = datetime.today().date()

# =====================================================================
# 1. PAGE INITIALIZATION & DATABASE CONFIGURATION (MUST BE FIRST)
# =====================================================================
st.set_page_config(
    page_title="Field Operations",
    layout="wide"
)

URL = st.secrets["SUPABASE_URL"]
KEY = st.secrets["SUPABASE_KEY"]


@st.cache_resource
def get_supabase_client() -> Client:
    return create_client(URL, KEY)


supabase = get_supabase_client()


# -----------------------------------------------------------------
# DYNAMIC USER REGISTRY LOADING
# -----------------------------------------------------------------
@st.cache_data(ttl=600)  # Caches for 10 minutes, clears on update
def load_user_registry():

    try:
        res = supabase.table("USER_REGISTRY").select("user_name").order("user_name", ascending=True).execute()
        if res.data and len(res.data) > 0:
            return ['----'] + [row['user_name'] for row in res.data]
        return ['----', 'ESP-KOC', 'JO-ESP', 'WORKSHOP', 'BURGUN-YRD', 'MOBILE', 'OFF-HIRE']  # Sensible fallback
    except Exception as e:
        return ['----', 'ESP-KOC', 'JO-ESP', 'WORKSHOP','BURGUN','MOBILE','OFF-HIRE','ABDALY-FARM','READY','DESALTER-PROJECT',
                'FIELD_OP.REPAIR','GAS-MITIGATION','MISHRIF','PDI','WS-POWER']  # Fault safe fallback


# Globally populated list used by all asset dropdown forms (DO NOT OVERWRITE BELOW)
USER_LIST = load_user_registry()


@st.cache_data(show_spinner=False)
def inject_custom_css(css_file_path: str = "style.css"):
    try:
        with open(css_file_path, "r") as f:
            css_style = f.read()
        return f"<style>{css_style}</style>"
    except FileNotFoundError:
        return ""


# --- DATABASE CACHING UTILITIES ---
@st.cache_data(ttl=15)
def get_assets_df() -> pd.DataFrame:
    try:
        res = supabase.table("ASSETS").select("*").order("id").execute()
        return pd.DataFrame(res.data)
    except Exception as e:
        st.error(f"Error connecting to assets database context: {e}")
        return pd.DataFrame()


def log_audit_event(g_code: str, action: str, status: str, description: str, old_data: dict = None,
                    new_data: dict = None):
    try:
        current_operator = st.session_state.get("auth_user_email", "SYSTEM_AUTO")
        log_payload = {
            "g_code": str(g_code).strip().upper(),
            "action_type": str(action).upper(),
            "transfer_status": str(status) if status and status != '----' else "N/A",
            "details": str(description),
            "old_values": old_data,
            "new_values": new_data,
            "changed_by": current_operator
        }
        supabase.table("AUDIT_LOGS").insert(log_payload).execute()
    except Exception as audit_err:
        st.error(f"⚠️ Log Insert Failure: {audit_err}")


def get_audit_logs_df() -> pd.DataFrame:
    try:
        res = supabase.table("AUDIT_LOGS").select("*").order("created_at", desc=True).execute()
        return pd.DataFrame(res.data) if res.data else pd.DataFrame()
    except Exception as e:
        st.error(f"Error fetching logs: {e}")
        return pd.DataFrame()


def delete_old_audit_logs(days_retention=30, purge_all=False):
    try:
        if purge_all:
            response = supabase.table("AUDIT_LOGS").delete().neq("g_code", "").execute()
        else:
            from datetime import datetime, timedelta, timezone
            cutoff_date = datetime.now(timezone.utc) - timedelta(days=days_retention)
            cutoff_iso = cutoff_date.isoformat()
            response = supabase.table("AUDIT_LOGS").delete().lt("created_at", cutoff_iso).execute()
        return {"status": "success", "row_count": len(response.data) if response.data else 0}
    except Exception as err:
        return {"status": "error", "message": str(err)}


#Application Base Matrix Profiles
TRANS_LIST = ['----', 'NEW GENERATOR', 'DISPATCH', 'RECEIVED', 'INTERNAL-SHIFT']


# Zebra Striping Engine Function Matrix
def style_zebra_rows(df):
    styles = pd.DataFrame('', index=df.index, columns=df.columns)
    styles.iloc[0::2, :] = 'background-color: #f0f9ff;'
    return styles


# =====================================================================
# 2. PERSISTENT SECURITY STATE INITIALIZATION KEYS
# =====================================================================
if "authenticated" not in st.session_state:
    st.session_state["authenticated"] = False
if "auth_user_email" not in st.session_state:
    st.session_state["auth_user_email"] = None
if "user_role" not in st.session_state:
    st.session_state["user_role"] = None

# =====================================================================
# 3. IDENTITY GATEWAY SECURITY ROUTING
# =====================================================================
if not st.session_state["authenticated"]:
    st.markdown(
        """
        <style>
            .stApp { background: linear-gradient(135deg, #f5f7fa 0%, #c3cfe2 100%) !important; }
            .block-container { padding-top: 5rem !important; }
            .login-card-container { background-color: #ffffff; border-radius: 14px; box-shadow: 0 8px 24px rgba(0,0,0,0.12); overflow: hidden; border: 1px solid #e0e4ec; }
            .login-blue-header { background: linear-gradient(90deg, #4da3ff, #80bfff); padding: 2rem 1.5rem; text-align: center; color: #ffffff !important; }
            .header-title-container { display: flex; align-items: center; justify-content: center; gap: 14px; margin-bottom: 0.4rem; }
            .header-logo { height: 45px; width: auto; object-fit: contain; }
            .login-blue-header h2 { color: #ffffff !important; margin: 0 !important; font-weight: 700 !important; font-size: 1.6rem !important; letter-spacing: 0.5px; line-height: 1.2; }
            .login-blue-header p { color: #f0f5ff !important; margin: 0 !important; font-size: 0.9rem !important; opacity: 0.9; padding-top: 0.5rem; }
            .login-form-body { padding: 2rem; }
        </style>
        """, unsafe_allow_html=True
    )


    def get_base64_image(img_path):
        if os.path.exists(img_path):
            with open(img_path, "rb") as img_file:
                return f"data:image/png;base64,{base64.b64encode(img_file.read()).decode()}"
        return "https://via.placeholder.com/150"


    logo_file_path = "logo.png"
    img_base64 = get_base64_image(logo_file_path)

    _, login_container_col, _ = st.columns([1, 1.8, 1])
    with login_container_col:
        st.markdown('<div class="login-card-container">', unsafe_allow_html=True)
        st.markdown(
            f"""
            <div class="login-blue-header">
                <div class="header-title-container">
                    <img src="{img_base64}" class="header-logo" alt="Corporate Logo">
                    <h2>Field Operations Digital Assets Monitoring System</h2>
                </div>
                <p>Enterprise Digital Asset Management Registry Identity Authentication Portal</p>
            </div>
            """, unsafe_allow_html=True
        )
        st.markdown('<div class="login-form-body">', unsafe_allow_html=True)

        with st.form(key="gateway_security_login_form"):
            email_input = st.text_input("Account Corporate Email Address:", placeholder="username@company.com").strip()
            password_input = st.text_input("Account Secret Credentials Key:", type="password", placeholder="••••••••")
            st.markdown("<br>", unsafe_allow_html=True)

            if st.form_submit_button("VALIDATE SECURITY CREDENTIALS", use_container_width=True):
                if not email_input or not password_input:
                    st.error("❌ Credentials validation failure: Input parameters cannot be blank.")
                else:
                    with st.spinner("Processing authorization handshakes..."):
                        try:
                            auth_res = supabase.auth.sign_in_with_password(
                                {"email": email_input, "password": password_input})
                            target_uid = auth_res.user.id
                            profile_query = supabase.table("user_profiles").select("role").eq("id",
                                                                                              target_uid).execute()

                            if profile_query.data:
                                user_assigned_role = profile_query.data[0]["role"]
                                st.session_state["authenticated"] = True
                                st.session_state["auth_user_email"] = auth_res.user.email
                                st.session_state["user_role"] = user_assigned_role
                                st.success(f"🎉 Welcome back! Role: {user_assigned_role}")
                                st.rerun()
                            else:
                                st.error(
                                    "❌ Access Denied: Your email is in Auth, but missing a row in 'user_profiles' table.")
                        except Exception as auth_fail:
                            st.error(f"❌ Authentication Denied: {str(auth_fail)}")
        st.markdown('</div></div>', unsafe_allow_html=True)

else:
    # =====================================================================
    # 4. DATA INITIALIZATION & ENVIRONMENT VIEW LAYOUT OVERLAYS (GATEWAY PROTECTED)
    # =====================================================================
    @st.cache_data(ttl=30)
    def get_mappings_df() -> pd.DataFrame:
        try:
            res = supabase.table("asset_mappings").select("*").order("type").execute()
            df = pd.DataFrame(res.data)
            if df.empty:
                return pd.DataFrame(columns=['id', 'type', 'model', 'kva'])
            return df
        except Exception as e:
            st.error(f"Error loading engineering model specification profiles: {e}")
            return pd.DataFrame(columns=['id', 'type', 'model', 'kva'])


    @st.cache_data(ttl=10)
    def get_routing_matrix_df() -> pd.DataFrame:
        try:
            res = supabase.table("field_routing_matrix").select("*").order("field_name").execute()
            df = pd.DataFrame(res.data)
            if df.empty:
                return pd.DataFrame(columns=['id', 'field_name', 'area_name', 'location_name'])
            return df
        except Exception as e:
            st.error(f"Error fetching structural routing profiles: {e}")
            return pd.DataFrame(columns=['id', 'field_name', 'area_name', 'location_name'])


    mappings_df = get_mappings_df()
    TYPE_LIST = ['----'] + sorted(mappings_df['type'].unique().tolist()) if not mappings_df.empty else ['----']
    routing_df = get_routing_matrix_df()
    LIVE_FIELD_OPTIONS = ['----'] + sorted(routing_df['field_name'].unique().tolist()) if not routing_df.empty else [
        '----']
    MAPPED_LOCATIONS_POOL = ['----'] + sorted(
        routing_df['location_name'].dropna().unique().tolist()) if not routing_df.empty else ['----']

    st.markdown(inject_custom_css("style.css"), unsafe_allow_html=True)
    st.markdown(
        """
        <style>
            header[data-testid="stHeader"] { background-color: rgba(0,0,0,0) !important; height: 3.5rem; z-index: 999; }
            .block-container { padding-top: 4.5rem !important; padding-bottom: 1rem; padding-left: 2rem; padding-right: 2rem; }
            .custom-topbar {
                position: fixed; top: 0; left: 0; right: 0; height: 50px; background-color: #1e3a8a;
                color: #ffffff; display: flex; align-items: center; justify-content: center;
                padding: 0 2rem; z-index: 9999; box-shadow: 0px 2px 5px rgba(0,0,0,0.2); font-family: sans-serif;
            }
            .topbar-brand { font-size: 1.2rem; font-weight: bold; letter-spacing: 1px; text-align: center; }
        </style>
        <div class="custom-topbar"><div class="topbar-brand">FIELD OPERATIONS DIGITAL ASSETS MONITORING SYSTEM</div></div>
        """, unsafe_allow_html=True
    )
    # Injecting Custom CSS to align tabs 2px below the blue header bar
    st.markdown("""
        <style>
        /* =====================================================================
           0. PULL TABS TO TOP (Reduces distance below blue bar to 2px)
           ===================================================================== */
        div[data-testid="stTabBar"] {
            margin-top: -40px !important; /* Pulls navigation tabs upwards */
            margin-bottom: 0px !important;
        }

        .element-container:has(div[data-baseweb="tab-list"]) {
            margin-top: -15px !important; /* Counteracts inner element wrapper gaps */
        }

        /* =====================================================================
           1. Global Tab Navigation Container Bar 
           ===================================================================== */
        div[data-baseweb="tab-list"] {
            background-color: #f8f9fa !important; /* Light subtle grey backdrop */
            border: 1px solid #e0e0e0 !important;   /* Clean outer boundary border */
            border-radius: 8px 8px 0px 0px !important; /* Rounded top edges */
            padding: 6px 12px 6px 12px !important;
            gap: 8px !important;                  /* Add separation space between tabs */
            margin-top: 2px !important;           /* Holds the absolute 2px boundary gap */
        }

        /* =====================================================================
           2. Inactive Tab Buttons
           ===================================================================== */
        button[data-baseweb="tab"] {
            background-color: #ffffff !important;
            border: 1px solid #dcdcdc !important;
            border-bottom: none !important; /* Keep bottom flat to merge with content frame */
            border-radius: 6px 6px 0px 0px !important;
            padding: 8px 16px !important;
            font-weight: 500 !important;
            color: #666666 !important;
            transition: all 0.2s ease-in-out !important;
        }

        /* 3. Hover state effect for tabs */
        button[data-baseweb="tab"]:hover {
            color: #1f77b4 !important; /* Subtle blue primary tint on hover */
            background-color: #f1f3f5 !important;
            border-color: #b0b0b0 !important;
        }

        /* =====================================================================
           4. Active / Selected Tab Button
           ===================================================================== */
        button[data-baseweb="tab"][aria-selected="true"] {
            background-color: #ffffff !important;
            color: #1f77b4 !important;      /* Accent color text */
            border-color: #1f77b4 !important; /* Accent color border line */
            font-weight: bold !important;
            box-shadow: 0px -2px 0px #1f77b4 inset !important; /* Thicker underline accent indicator */
        }

        /* =====================================================================
           5. Content Boundary Card Frame Beneath Tabs
           ===================================================================== */
        div[data-testid="stTab"] {
            background-color: #ffffff !important;
            border: 1px solid #e0e0e0 !important;
            border-top: none !important;               /* Merge directly with the tab list bar above */
            border-radius: 0px 0px 8px 8px !important; /* Round out the bottom chassis */
            padding: 24px !important;                  /* Clean breathing room for fields inside */
            box-shadow: 0px 4px 6px rgba(0, 0, 0, 0.03) !important; /* Soft modern elevation drop shadow */
        }
        </style>
    """, unsafe_allow_html=True)

    user_clearance_role = st.session_state["user_role"]
    if user_clearance_role in ["MANAGER", "DEVELOPER", "SUPERVISOR", "ENGINEER"]:
        allowed_views = ["GENERAL_ASSETS", "ASSET MANAGEMENT", "AUDIT LOGS", "WORKSHOP", "STORES & PARTS",
                         "MAINTENANCE", "FLEET MANAGEMENT", "REMOTE TELEMETRY", "REPORTS", "SETTINGS"]
    else:
        allowed_views = ["GENERAL_ASSETS", "ASSET MANAGEMENT"]

    # =====================================================================
    # 5. SIDEBAR CONTROLLER WRAPPER
    # =====================================================================
    st.sidebar.title("GENSET FIELD_OPERATIONS")
    st.sidebar.markdown(
        f"👤 **Identity:** `{st.session_state['auth_user_email']}`  \n🛡️ **Role:** `{user_clearance_role}`")

    navigation_target = st.sidebar.radio("Choose an operational viewport module:", options=allowed_views,
                                         label_visibility="visible")

    st.sidebar.divider()
    if st.sidebar.button("🔄 Refresh page", use_container_width=True):
        st.cache_data.clear()
        st.rerun()

    if st.sidebar.button("🚪 Log-out", use_container_width=True):
        try:
            supabase.auth.sign_out()
        except:
            pass
        st.session_state["authenticated"] = False
        st.session_state["auth_user_email"] = None
        st.session_state["user_role"] = None
        st.rerun()

    # =====================================================================
    # 1. GENERAL_ASSETS RUNTIME VIEWPORT (UPGRADED WITH POWER BI FEATURES)
    # =====================================================================
    if navigation_target == "GENERAL_ASSETS":
        df = get_assets_df()
        if not df.empty:
            try:
                import plotly.express as px
                import pandas as pd

                # --- CRITICAL: SANITIZE DATA TYPES TO PREVENT RUNTIME ERRORS ---
                if 'GEN_KVA' in df.columns:
                    df['GEN_KVA'] = pd.to_numeric(df['GEN_KVA'], errors='coerce').fillna(0).astype(int)

                if 'RUN_HRS' in df.columns:
                    df['RUN_HRS'] = pd.to_numeric(df['RUN_HRS'], errors='coerce').fillna(0)

                # --- POWER BI FEATURE 1: DYNAMIC SIDEBAR FILTER SLICERS ---
                st.sidebar.markdown("---")
                st.sidebar.markdown("### 📊 Dashboard Slicers")

                # Slicer 1: Generator Type Selector
                available_types = ['ALL'] + sorted(df['TYPE'].dropna().unique().tolist()) if 'TYPE' in df.columns else [
                    'ALL']
                selected_type = st.sidebar.selectbox("Filter by Generator Type:", options=available_types,
                                                     key="bi_slicer_type")

                # Slicer 2: KVA Rating Range Slider
                if 'GEN_KVA' in df.columns and len(df) > 0:
                    min_kva_val = int(df['GEN_KVA'].min())
                    max_kva_val = int(df['GEN_KVA'].max())
                    if min_kva_val == max_kva_val:
                        max_kva_val = min_kva_val + 10
                    selected_kva_range = st.sidebar.slider("Filter by KVA Power Range:", min_kva_val, max_kva_val,
                                                           (min_kva_val, max_kva_val), key="bi_slicer_kva")
                else:
                    selected_kva_range = (0, 10000)

                # --- APPLY CROSS-FILTERING TO DATAFRAME DYNAMICALLY ---
                if selected_type != 'ALL':
                    df = df[df['TYPE'] == selected_type]

                if 'GEN_KVA' in df.columns:
                    df = df[(df['GEN_KVA'] >= selected_kva_range[0]) & (df['GEN_KVA'] <= selected_kva_range[1])]

                # --- RE-CALCULATE METRICS BASED ON FILTERED DATAFRAME ---
                user_counts = df['USER'].value_counts() if 'USER' in df.columns else pd.Series()
                v_ESP = user_counts.get('ESP-KOC', 0)
                v_WHP = user_counts.get('WORKSHOP', 0)
                v_JO = user_counts.get('JO-ESP', 0)
                v_BYRD = user_counts.get('BURGUN-YRD', 0)
                v_MBL = user_counts.get('MOBILE', 0)
                v_OFF = user_counts.get('OFF-HIRE', 0)
                v_abd = user_counts.get('ABDALY-FARM', 0)
                v_RDY = user_counts.get('READY', 0)
                v_DST = user_counts.get('DESALTER-PROJECT', 0)
                v_FD = user_counts.get('FIELD_OP.REPAIR', 0)
                v_GAS = user_counts.get('GAS-MITIGATION', 0)
                v_MHF = user_counts.get('MISHRIF', 0)
                v_PDI = user_counts.get('PDI', 0)
                v_WPP = user_counts.get('WS-POWER', 0)
                v_TT = len(df)

                # --- POWER BI FEATURE 2: HIGH-LEVEL EXECUTIVE SUMMARY KPI CARDS ---
                total_kva = df['GEN_KVA'].sum() if 'GEN_KVA' in df.columns else 0
                avg_runtime = df['RUN_HRS'].mean() if 'RUN_HRS' in df.columns else 0

                st.caption("📊 Quick Fleet Metrics Analysis")
                kpi1, kpi2, kpi3 = st.columns(3)
                with kpi1:
                    st.metric(label="📟 TOTAL ACTIVE FLEET", value=f"{v_TT:,} Units", border=True)
                with kpi2:
                    st.metric(label="⚡ TOTAL ACCUMULATED CAPACITY", value=f"{total_kva:,} KVA", border=True)
                with kpi3:
                    st.metric(label="⏳ AVG FLEET RUNNING HOURS",
                              value=f"{int(avg_runtime):,} Hrs" if pd.notna(avg_runtime) else "0 Hrs", border=True)

                st.markdown("---")

                # --- 💡 MOVED UP: LIVE RECORD VIEWPORT METRICS BREAKDOWN ---
                st.markdown("##### 📋 Live Filtered Record Viewport Summary")
                col1, col2, col3, col4, col5 = st.columns(5)
                with col1:
                    st.metric('ESP-KOC', int(v_ESP))
                    st.metric('WORKSHOP', int(v_WHP))
                    st.metric('JO-ESP', int(v_JO))
                with col2:
                    st.metric('BURGUN-YRD', int(v_BYRD))
                    st.metric('MOBILE', int(v_MBL))
                    st.metric('OFF-HIRE', int(v_OFF))
                with col3:
                    st.metric('ABDALY-FARM', int(v_abd))
                    st.metric('READY', int(v_RDY))
                    st.metric('DESALTER-PROJECT', int(v_DST))
                with col4:
                    st.metric('FIELD_OP.REPAIR', int(v_FD))
                    st.metric('GAS-MITIGATION', int(v_GAS))
                    st.metric('MISHRIF', int(v_MHF))
                with col5:
                    st.metric('PDI', int(v_PDI))
                    st.metric('WS-POWER', int(v_WPP))
                    st.metric('FILTERED TOTAL', int(v_TT))

                st.markdown("---")

                # --- POWER BI FEATURE 3: SIDE-BY-SIDE INTERACTIVE VISUALS ---
                chart_col1, chart_col2 = st.columns([2, 3])

                with chart_col1:
                    st.markdown("##### 🍩 Fleet Distribution (USER / STATUS)")
                    if v_TT > 0:
                        status_map = {
                            'ESP-KOC': v_ESP, 'WORKSHOP': v_WHP, 'JO-ESP': v_JO,
                            'BURGUN-YRD': v_BYRD, 'MOBILE': v_MBL, 'OFF-HIRE': v_OFF,
                            'ABDALY-FARM': v_abd, 'READY': v_RDY, 'DESALTER-PROJECT': v_DST,
                            'FIELD_OP.REPAIR': v_FD, 'GAS-MITIGATION': v_GAS, 'MISHRIF': v_MHF,
                            'PDI': v_PDI, 'WS-POWER': v_WPP
                        }
                        chart_df = pd.DataFrame(list(status_map.items()), columns=['Operational Group', 'Count'])
                        chart_df = chart_df[chart_df['Count'] > 0]

                        if not chart_df.empty:
                            fig_donut = px.pie(
                                chart_df,
                                names='Operational Group',
                                values='Count',
                                hole=0.45,
                                color_discrete_sequence=px.colors.qualitative.Pastel
                            )
                            fig_donut.update_layout(margin=dict(t=10, b=10, l=10, r=10), height=320, showlegend=True)
                            st.plotly_chart(fig_donut, use_container_width=True)
                        else:
                            st.info("No active units split within selection categories.")
                    else:
                        st.info("No data available within selection filters.")

                with chart_col2:
                    st.markdown("##### 📊 Capacity Allocation Matrix By Field Location")
                    if v_TT > 0 and 'FIELD' in df.columns and 'GEN_KVA' in df.columns:
                        field_kva_df = df.groupby('FIELD')['GEN_KVA'].sum().reset_index()
                        field_kva_df = field_kva_df[field_kva_df['FIELD'] != '----']
                        field_kva_df = field_kva_df.sort_values(by='GEN_KVA', ascending=True)

                        fig_bar = px.bar(
                            field_kva_df,
                            x='GEN_KVA',
                            y='FIELD',
                            orientation='h',
                            color='GEN_KVA',
                            color_continuous_scale='Blues',
                            labels={'GEN_KVA': 'Total Capacity (KVA)', 'FIELD': 'Field Assignment'}
                        )
                        fig_bar.update_layout(margin=dict(t=10, b=10, l=10, r=10), height=320,
                                              coloraxis_showscale=False)
                        st.plotly_chart(fig_bar, use_container_width=True)
                    else:
                        st.info("No field assignments match filtering constraints.")

            except Exception as e:
                st.error(f"Error parsing status metrics: {e}")
        else:
            st.info("No equipment inventory assets found inside database registries.")


    # =====================================================================
    # 2. ASSET MANAGEMENT ACTIONS ENGINE
    # =====================================================================
    elif navigation_target == "ASSET MANAGEMENT":
        df = get_assets_df()
        tab1, tab2, tab3, tab4, tab5 ,tab6= st.tabs(
            ['ASSETS_VIEW', 'ADD_ASSET', 'UPDATE_ASSET', 'DELETE_ASSET', 'AUDIT LOGS','ASSET-SPECTS'])

        with tab1:
            if not df.empty:
                designed_assets_df = df.style.apply(style_zebra_rows, axis=None)
                st.dataframe(designed_assets_df, use_container_width=True, hide_index=True,height=700)
            else:
                st.info("No equipment inventory assets found inside database registries.")

        with tab2:
            st.caption("ADDING NEW ASSET",text_alignment="center")
            m_col1, m_col2, m_col3 = st.columns(3)
            with m_col1:
                u_type = st.selectbox('Select Manufacturer TYPE:', options=TYPE_LIST, key="add_type_outside")
            with m_col2:
                add_filtered = mappings_df[mappings_df['type'] == u_type] if u_type != '----' else pd.DataFrame()
                add_allowed_models = ['----'] + sorted(
                    add_filtered['model'].unique().tolist()) if not add_filtered.empty else ['----']
                u_model = st.selectbox('Select Engine MODEL:', options=add_allowed_models, key="add_model_outside")
            with m_col3:
                add_matched_row = add_filtered[
                    add_filtered['model'] == u_model] if u_model != '----' else pd.DataFrame()
                calculated_kva = int(add_matched_row.iloc[0]['kva']) if not add_matched_row.empty else 0
                u_kva = st.number_input('Assigned Rating (KVA):', min_value=0, value=calculated_kva, step=10,
                                        key="add_kva_outside")

            cascade_col1, cascade_col2, cascade_col3 = st.columns(3)
            with cascade_col1:
                u_field = st.selectbox('TO_FIELD :', options=LIVE_FIELD_OPTIONS, key="add_field_cascade")
            with cascade_col2:
                field_matched_df = routing_df[
                    routing_df['field_name'] == u_field] if u_field != '----' else pd.DataFrame()
                ALLOWED_AREAS = ['----'] + sorted(
                    field_matched_df['area_name'].unique().tolist()) if not field_matched_df.empty else ['----']
                u_area = st.selectbox('AREA :', options=ALLOWED_AREAS, key="add_area_cascade")
            with cascade_col3:
                area_matched_df = field_matched_df[
                    field_matched_df['area_name'] == u_area] if u_area != '----' else pd.DataFrame()
                ALLOWED_LOCATIONS = ['----'] + sorted(
                    area_matched_df['location_name'].unique().tolist()) if not area_matched_df.empty else ['----']
                u_location = st.selectbox('TO_LOCATION :', options=ALLOWED_LOCATIONS, key="add_location_cascade")

            # --- CSS Mixin to adjust vertical text-align balance ---
            LABEL_STYLE = "<p style='margin-top:8px; font-weight:bold; text-align:right; font-size:13px; color:#333333;'>{}</p>"

            with st.form('ASSET_ADD_LOGISTICS_FORM', clear_on_submit=True):


                # --- ROW 1 ---
                r1_lbl1, r1_val1, r1_lbl2, r1_val2, r1_lbl3, r1_val3 = st.columns([1.2, 2, 1.2, 2, 1.2, 2])
                with r1_lbl1:
                    st.markdown(LABEL_STYLE.format("TRANSFER:"), unsafe_allow_html=True)
                with r1_val1:
                    u_transfer = st.selectbox('', options=TRANS_LIST, key="add_transfer", label_visibility="collapsed")
                with r1_lbl2:
                    st.markdown(LABEL_STYLE.format("SERIAL NO:"), unsafe_allow_html=True)
                with r1_val2:
                    u_serial = st.text_input('', value='', key="add_serial", label_visibility="collapsed")
                with r1_lbl3:
                    st.markdown(LABEL_STYLE.format("RUN HOURS:"), unsafe_allow_html=True)
                with r1_val3:
                    u_run_hr = st.number_input('', min_value=0, key="add_run_hrs", label_visibility="collapsed")

                # --- ROW 2 ---
                r2_lbl1, r2_val1, r2_lbl2, r2_val2, r2_lbl3, r2_val3 = st.columns([1.2, 2, 1.2, 2, 1.2, 2])
                with r2_lbl1:
                    st.markdown(LABEL_STYLE.format("FROM LOC:"), unsafe_allow_html=True)
                with r2_val1:
                    u_from_location = st.selectbox('', options=MAPPED_LOCATIONS_POOL, key="add_from_loc",
                                                   label_visibility="collapsed")
                with r2_lbl2:
                    st.markdown(LABEL_STYLE.format("MANUF YR:"), unsafe_allow_html=True)
                with r2_val2:
                    u_manuf_date = st.date_input('', min_value=min_date, max_value=max_date, key="add_manuf",
                                                 label_visibility="collapsed")
                with r2_lbl3:
                    st.markdown(LABEL_STYLE.format("APPR KVA:"), unsafe_allow_html=True)
                with r2_val3:
                    u_appr_kva = st.number_input('', min_value=0, key="add_appr", label_visibility="collapsed")

                # --- ROW 3 ---
                r3_lbl1, r3_val1, r3_lbl2, r3_val2, r3_lbl3, r3_val3 = st.columns([1.2, 2, 1.2, 2, 1.2, 2])
                with r3_lbl1:
                    st.markdown(LABEL_STYLE.format("G-CODE:"), unsafe_allow_html=True)
                with r3_val1:
                    g_code = st.text_input('', value='', key="add_gcode", label_visibility="collapsed")
                with r3_lbl2:
                    st.markdown(LABEL_STYLE.format("SERVICE YR:"), unsafe_allow_html=True)
                with r3_val2:
                    u_service_yr = st.date_input('', min_value=min_date, max_value=max_date, key="add_service",
                                                 label_visibility="collapsed")
                with r3_lbl3:
                    st.markdown(LABEL_STYLE.format("USER:"), unsafe_allow_html=True)
                with r3_val3:
                    u_user = st.selectbox('', options=USER_LIST, key="add_user", label_visibility="collapsed")

                # --- ROW 4 ---
                r4_lbl1, r4_val1, r4_lbl2, r4_val2 = st.columns([1.2, 2, 1.2, 5.2])
                with r4_lbl1:
                    st.markdown(LABEL_STYLE.format("MOVE DATE:"), unsafe_allow_html=True)
                with r4_val1:
                    u_move_date = st.date_input('', min_value=min_date, max_value=max_date, key="add_move_dt",
                                                label_visibility="collapsed")
                with r4_lbl2:
                    st.markdown(LABEL_STYLE.format("REMARKS:"), unsafe_allow_html=True)
                with r4_val2:
                    u_reason = st.text_area("", value='', key="add_reason", height=42, label_visibility="collapsed")

                if st.form_submit_button("SAVE NEW ASSET TO DATABASE", use_container_width=True):
                    if not g_code.strip():
                        st.error("❌ Database validation rejected: G-CODE field cannot be blank.")
                    else:
                        new_data = {
                            'TRANSFER_STATUS': u_transfer if u_transfer != '----' else None,
                            'FIELD': u_field if u_field != '----' else None,
                            'AREA': u_area if u_area != '----' else None,
                            'TO_LOCATION': u_location if u_location != '----' else None,
                            'FROM_LOCATION': u_from_location if u_from_location != '----' else None,
                            'G-CODE': g_code.strip().upper(),
                            'SERIAL_NO': u_serial.strip().upper(),
                            'MODEL': u_model if u_model != '----' else None,
                            'TYPE': u_type if u_type != '----' else None,
                            'GEN_KVA': u_kva,
                            'MANUF_YR': u_manuf_date.isoformat(),
                            'KOC_SERVICE_YR': u_service_yr.isoformat(),
                            'RUN_HRS': u_run_hr,
                            'APPR_KVA': u_appr_kva,
                            'USER': u_user if u_user != '----' else None,
                            'MOVE_DATE': u_move_date.isoformat(),
                            'REASON': u_reason,
                        }
                        try:
                            supabase.table("ASSETS").insert(new_data).execute()
                            st.success(f"🎉 Asset '{g_code.upper()}' recorded successfully!")
                            st.cache_data.clear()
                            st.rerun()
                        except Exception as err:
                            st.error(f"Supabase Transmission Error: {err}")

        # =====================================================================
        # TAB 3: UPDATE / MODIFY EXISTING FIELD ASSET
        # =====================================================================
        with tab3:
            st.caption("Modify Existing Field Asset",text_alignment="center")
            if not df.empty:
                asset_options = sorted(df['G-CODE'].dropna().unique().tolist())
                selected_gcode = st.selectbox("Select Asset G-CODE to Update:", options=asset_options,
                                              key="select_gcode_updater")

                with st.spinner(f"Retrieving live record for {selected_gcode}..."):
                    try:
                        live_res = supabase.table("ASSETS").select("*").eq("G-CODE", selected_gcode).execute()
                        if live_res.data:
                            asset_row = live_res.data[0]
                        else:
                            st.error(f"Asset {selected_gcode} could not be located.")
                            st.stop()
                    except Exception as db_err:
                        st.error(f"Failed to fetch live asset data: {db_err}")
                        st.stop()


                def get_index(opt_list, val):
                    v = str(val).strip() if pd.notna(val) and val is not None else '----'
                    return opt_list.index(v) if v in opt_list else 0

                up_cascade_col1, up_cascade_col2, up_cascade_col3 = st.columns(3)

                with up_cascade_col1:
                    db_field = asset_row.get('FIELD', '----')
                    up_field = st.selectbox('TO_FIELD :', options=LIVE_FIELD_OPTIONS,
                                            index=get_index(LIVE_FIELD_OPTIONS, db_field),
                                            key=f"up_field_cascade_{selected_gcode}")
                with up_cascade_col2:
                    up_field_matched_df = routing_df[
                        routing_df['field_name'] == up_field] if up_field != '----' else pd.DataFrame()
                    UP_ALLOWED_AREAS = ['----'] + sorted(
                        up_field_matched_df['area_name'].unique().tolist()) if not up_field_matched_df.empty else [
                        '----']
                    db_area = asset_row.get('AREA', '----')
                    up_area = st.selectbox('TO_AREA :', options=UP_ALLOWED_AREAS,
                                           index=get_index(UP_ALLOWED_AREAS, db_area),
                                           key=f"up_area_cascade_{selected_gcode}")
                with up_cascade_col3:
                    up_area_matched_df = up_field_matched_df[
                        up_field_matched_df['area_name'] == up_area] if up_area != '----' else pd.DataFrame()
                    UP_ALLOWED_LOCATIONS = ['----'] + sorted(
                        up_area_matched_df['location_name'].unique().tolist()) if not up_area_matched_df.empty else [
                        '----']
                    db_loc = asset_row.get('TO_LOCATION', '----')
                    up_location = st.selectbox('TO_LOCATION :', options=UP_ALLOWED_LOCATIONS,
                                               index=get_index(UP_ALLOWED_LOCATIONS, db_loc),
                                               key=f"up_to_loc_cascade_{selected_gcode}")

                allow_submission = True
                if up_location != '----' and up_location != db_loc:
                    try:
                        existing_check = supabase.table("ASSETS").select("G-CODE", "MODEL", "TRANSFER_STATUS").eq(
                            "TO_LOCATION", up_location).neq("G-CODE", selected_gcode).execute()

                        if existing_check.data:
                            allow_submission = False
                            clashing_assets = ", ".join(
                                [f"{item.get('G-CODE')} ({item.get('MODEL', 'Unknown Model')})" for item in
                                 existing_check.data])
                            st.error(
                                f"⚠️ **LOCATION CONFLICT:** The location **{up_location}** already has an active asset assigned to it: **{clashing_assets}**.")

                            bypass_checkbox = st.checkbox(
                                f"🚨 Allow secondary assignment? Check this box if want to deploy multiple assets to {up_location}.",
                                key=f"bypass_conflict_{selected_gcode}"
                            )
                            if bypass_checkbox:
                                allow_submission = True
                                st.success("🔓 Secondary deployment authorized by user.")
                    except Exception as check_err:
                        st.warning(f"Unable to verify location history entries: {check_err}")

                st.markdown("---")
                lock_engine_specs = asset_row.get('TRANSFER_STATUS') in ["DISPATCH", "RECEIVED", "INTERNAL-SHIFT"]

                db_type = asset_row.get('TYPE', '----')
                db_model = asset_row.get('MODEL', '----')

                try:
                    db_kva = int(float(asset_row.get('GEN_KVA', 0)))
                except:
                    db_kva = 0

                type_key = f"form_type_{selected_gcode}"
                model_key = f"form_model_{selected_gcode}"

                if type_key not in st.session_state:
                    st.session_state[type_key] = db_type if db_type in TYPE_LIST else '----'

                current_type = st.session_state[type_key]
                up_filtered = mappings_df[
                    mappings_df['type'] == current_type] if current_type != '----' else pd.DataFrame()
                up_allowed_models = ['----'] + sorted(
                    up_filtered['model'].unique().tolist()) if not up_filtered.empty else ['----']

                if model_key not in st.session_state:
                    st.session_state[model_key] = db_model if (
                                current_type == db_type and db_model in up_allowed_models) else '----'

                with st.form(key=f"ASSET_UPDATE_FORM_GROUP_{selected_gcode}"):
                    cfg_lbl1, cfg_val1, cfg_lbl2, cfg_val2, cfg_lbl3, cfg_val3 = st.columns([1.2, 2, 1.2, 2, 1.2, 2])
                    with cfg_lbl1:
                        st.markdown(LABEL_STYLE.format("TYPE:"), unsafe_allow_html=True)
                    with cfg_val1:
                        up_type = st.selectbox('', options=TYPE_LIST, index=TYPE_LIST.index(st.session_state[type_key]),
                                               key=type_key, disabled=lock_engine_specs, label_visibility="collapsed")
                    with cfg_lbl2:
                        st.markdown(LABEL_STYLE.format("MODEL:"), unsafe_allow_html=True)
                    with cfg_val2:
                        model_idx = up_allowed_models.index(st.session_state[model_key]) if st.session_state[
                                                                                                model_key] in up_allowed_models else 0
                        up_model = st.selectbox('', options=up_allowed_models, index=model_idx, key=model_key,
                                                disabled=lock_engine_specs, label_visibility="collapsed")
                    with cfg_lbl3:
                        st.markdown(LABEL_STYLE.format("KVA RATING:"), unsafe_allow_html=True)
                    with cfg_val3:
                        up_matched = up_filtered[
                            up_filtered['model'] == up_model] if up_model != '----' else pd.DataFrame()
                        fallback_kva = int(up_matched.iloc[0]['kva']) if not up_matched.empty else 0
                        initial_kva = db_kva if (up_type == db_type and up_model == db_model) else fallback_kva
                        up_kva = st.number_input('', min_value=0, value=initial_kva, step=10,
                                                 key=f"up_kva_widget_{selected_gcode}", disabled=lock_engine_specs,
                                                 label_visibility="collapsed")


                    # --- ROW 1 ---
                    u1_lbl1, u1_val1, u1_lbl2, u1_val2, u1_lbl3, u1_val3 = st.columns([1.2, 2, 1.2, 2, 1.2, 2])
                    with u1_lbl1:
                        st.markdown(LABEL_STYLE.format("TRANSFER:"), unsafe_allow_html=True)
                    with u1_val1:
                        up_transfer = st.selectbox('', options=TRANS_LIST,
                                                   index=get_index(TRANS_LIST, asset_row.get('TRANSFER_STATUS')),
                                                   key=f"up_trans_status_{selected_gcode}",
                                                   label_visibility="collapsed")
                    with u1_lbl2:
                        st.markdown(LABEL_STYLE.format("FROM LOC:"), unsafe_allow_html=True)
                    with u1_val2:
                        up_from_location = st.selectbox('', options=MAPPED_LOCATIONS_POOL,
                                                        index=get_index(MAPPED_LOCATIONS_POOL,
                                                                        asset_row.get('FROM_LOCATION')),
                                                        key=f"up_from_loc_{selected_gcode}",
                                                        label_visibility="collapsed")
                    with u1_lbl3:
                        st.markdown(LABEL_STYLE.format("MANUF YR:"), unsafe_allow_html=True)
                    with u1_val3:
                        def safe_date(val):
                            import datetime as dt_mod
                            if pd.isna(val) or not val or str(val).strip() in ["—", "----"]: return dt_mod.date.today()
                            if isinstance(val, (dt_mod.datetime, dt_mod.date)): return val if isinstance(val,
                                                                                                         dt_mod.date) else val.date()
                            try:
                                return dt_mod.datetime.strptime(str(val).split()[0], "%Y-%m-%d").date()
                            except:
                                return dt_mod.date.today()


                        up_manuf_date = st.date_input('', value=safe_date(asset_row.get('MANUF_YR')),
                                                      key=f"up_manuf_{selected_gcode}", disabled=lock_engine_specs,
                                                      label_visibility="collapsed")

                    # --- ROW 2 ---
                    u2_lbl1, u2_val1, u2_lbl2, u2_val2, u2_lbl3, u2_val3 = st.columns([1.2, 2, 1.2, 2, 1.2, 2])
                    with u2_lbl1:
                        st.markdown(LABEL_STYLE.format("USER:"), unsafe_allow_html=True)
                    with u2_val1:
                        up_user = st.selectbox('', options=USER_LIST, index=get_index(USER_LIST, asset_row.get('USER')),
                                               key=f"up_user_{selected_gcode}", label_visibility="collapsed")
                    with u2_lbl2:
                        st.markdown(LABEL_STYLE.format("SERIAL NO:"), unsafe_allow_html=True)
                    with u2_val2:
                        up_serial = st.text_input('', value=str(asset_row.get('SERIAL_NO', '')) if asset_row.get(
                            'SERIAL_NO') is not None else '', key=f"up_serial_{selected_gcode}",
                                                  disabled=lock_engine_specs, label_visibility="collapsed")
                    with u2_lbl3:
                        st.markdown(LABEL_STYLE.format("SERVICE YR:"), unsafe_allow_html=True)
                    with u2_val3:
                        up_service_yr = st.date_input('', value=safe_date(asset_row.get('KOC_SERVICE_YR')),
                                                      key=f"up_service_{selected_gcode}", disabled=lock_engine_specs,
                                                      label_visibility="collapsed")

                    # --- ROW 3 ---
                    u3_lbl1, u3_val1, u3_lbl2, u2_val2, u3_lbl3, u3_val3 = st.columns([1.2, 2, 1.2, 2, 1.2, 2])
                    with u3_lbl1:
                        st.markdown(LABEL_STYLE.format("PURPOSE:"), unsafe_allow_html=True)
                    with u3_val1:
                        PURPOSE_LIST = ['----', 'REPLACEMENT', 'NEW INSTALLATION', 'OFF-HIRE BACKLOAD']
                        db_purpose = str(asset_row.get('PURPOSE', '')).strip().upper()
                        up_purpose = st.selectbox('', options=PURPOSE_LIST, index=PURPOSE_LIST.index(
                            db_purpose) if db_purpose in PURPOSE_LIST else 0, key=f"up_purpose_{selected_gcode}",
                                                  label_visibility="collapsed")
                    with u3_lbl2:
                        st.markdown(LABEL_STYLE.format("APPR KVA:"), unsafe_allow_html=True)
                    with u2_val2:
                        try:
                            current_appr_kva = int(float(asset_row.get('APPR_KVA', 0)))
                        except:
                            current_appr_kva = 0
                        up_appr_kva = st.number_input('', min_value=0, value=current_appr_kva,
                                                      key=f"up_appr_{selected_gcode}", label_visibility="collapsed")
                    with u3_lbl3:
                        st.markdown(LABEL_STYLE.format("RUN HOURS:"), unsafe_allow_html=True)
                    with u3_val3:
                        try:
                            current_run_hrs = int(float(asset_row.get('RUN_HRS', 0)))
                        except:
                            current_run_hrs = 0
                        up_run_hr = st.number_input('', min_value=0, value=current_run_hrs,
                                                    key=f"up_run_hrs_{selected_gcode}", label_visibility="collapsed")

                    # --- ROW 4 ---
                    u4_lbl1, u4_val1, u4_lbl2, u4_val2, u4_lbl3, u4_val3 = st.columns([1.2, 2, 1.2, 2, 1.2, 2])
                    with u4_lbl1:
                        st.markdown(LABEL_STYLE.format("CREW:"), unsafe_allow_html=True)
                    with u4_val1:
                        try:
                            current_crew = int(float(asset_row.get('CREW', 0)))
                        except:
                            current_crew = 0
                        up_crew = st.number_input('', min_value=0, value=current_crew, step=1,
                                                  key=f"up_crew_{selected_gcode}", label_visibility="collapsed")
                    with u4_lbl2:
                        st.markdown(LABEL_STYLE.format("GC FIELDS:"), unsafe_allow_html=True)
                    with u4_val2:
                        try:
                            current_gc = int(float(asset_row.get('GC', 0)))
                        except:
                            current_gc = 0
                        up_gc = st.number_input('', min_value=0, value=current_gc, step=1,
                                                key=f"up_gc_{selected_gcode}", label_visibility="collapsed")
                    with u4_lbl3:
                        st.markdown(LABEL_STYLE.format("MOVE DATE:"), unsafe_allow_html=True)
                    with u4_val3:
                        up_move_date = st.date_input('', value=safe_date(asset_row.get('MOVE_DATE')),
                                                     key=f"up_move_dt_{selected_gcode}", label_visibility="collapsed")



                    # --- ROW 5 (Remarks) ---
                    rem_lbl, rem_val = st.columns([1.2, 8.4])
                    with rem_lbl:
                        st.markdown(LABEL_STYLE.format("REMARKS:"), unsafe_allow_html=True)
                    with rem_val:
                        up_reason = st.text_area("", value=str(asset_row.get('REASON', '')) if asset_row.get(
                            'REASON') is not None else '', key=f"up_reason_{selected_gcode}",
                                                 label_visibility="collapsed", height=52)

                    if st.form_submit_button("CLICK TO UPDATE ASSET", use_container_width=True,
                                             key=f"up_btn_{selected_gcode}"):
                        if not allow_submission:
                            st.error(
                                "❌ Action Blocked: You must resolve or acknowledge the Location Conflict before saving changes.")
                        elif up_run_hr < current_run_hrs:
                            st.error(
                                f"❌ Updated hours ({up_run_hr:,}) cannot run lower than current data entries ({current_run_hrs:,} hrs).")
                        else:
                            try:
                                before_q = supabase.table("ASSETS").select("*").eq('G-CODE', selected_gcode).execute()
                                old_snapshot = before_q.data[0] if before_q.data else {}
                            except:
                                old_snapshot = {}

                            updated_payload = {
                                'TRANSFER_STATUS': up_transfer if up_transfer != '----' else old_snapshot.get(
                                    'TRANSFER_STATUS'),
                                'FIELD': up_field if up_field != '----' else old_snapshot.get('FIELD'),
                                'AREA': up_area if up_area != '----' else old_snapshot.get('AREA'),
                                'TO_LOCATION': up_location if up_location != '----' else old_snapshot.get(
                                    'TO_LOCATION'),
                                'FROM_LOCATION': up_from_location if up_from_location != '----' else old_snapshot.get(
                                    'FROM_LOCATION'),
                                'SERIAL_NO': up_serial,
                                'MODEL': up_model if up_model != '----' else old_snapshot.get('MODEL'),
                                'TYPE': up_type if up_type != '----' else old_snapshot.get('TYPE'),
                                'GEN_KVA': up_kva,
                                'MANUF_YR': up_manuf_date.isoformat(),
                                'KOC_SERVICE_YR': up_service_yr.isoformat(),
                                'RUN_HRS': up_run_hr,
                                'APPR_KVA': up_appr_kva,
                                'USER': up_user if up_user != '----' else old_snapshot.get('USER'),
                                'MOVE_DATE': up_move_date.isoformat(),
                                'REASON': up_reason,
                                'PURPOSE': up_purpose if up_purpose != '----' else None,
                                'CREW': up_crew,
                                'GC': up_gc
                            }

                            try:
                                supabase.table("ASSETS").update(updated_payload).eq('G-CODE', selected_gcode).execute()
                                after_q = supabase.table("ASSETS").select("*").eq('G-CODE', selected_gcode).execute()
                                new_snapshot = after_q.data[0] if after_q.data else updated_payload

                                log_audit_event(selected_gcode, "UPDATE", str(up_transfer),
                                                f"Advanced runtime parameter metrics: {current_run_hrs:,} -> {up_run_hr:,} hrs.",
                                                old_snapshot, new_snapshot)

                                st.success("Asset successfully synchronized across network registries!")
                                st.cache_data.clear()
                                st.rerun()
                            except Exception as update_err:
                                st.error(f"Supabase Execution Sync Fault: {update_err}")
            else:
                st.info("No master record data currently loaded.")

        with tab4:
            if not df.empty:
                # 1. Provide an isolated drop-down to pick the asset to delete
                delete_options = sorted(df['G-CODE'].dropna().unique().tolist())
                target_gcode = st.selectbox(
                    "Select Asset G-CODE to PERMANENTLY Delete:",
                    options=delete_options,
                    key="select_gcode_deletion_mgr"
                )

                # 2. Retrieve a quick snapshot of the record before dropping it (for the audit trail)
                with st.spinner(f"Verifying existence of {target_gcode}..."):
                    try:
                        live_check = supabase.table("ASSETS").select("*").eq("G-CODE", target_gcode).execute()
                        if live_check.data:
                            deletion_snapshot = live_check.data[0]

                            # Display a quick summary so the operator knows exactly what they are deleting
                            st.info(
                                f"**Asset Identified:** {target_gcode}  \n"
                                f"**Type/Model:** {deletion_snapshot.get('TYPE', '—')} / {deletion_snapshot.get('MODEL', '—')}  \n"
                                f"**Current Location:** {deletion_snapshot.get('TO_LOCATION', '—')}  \n"
                                f"**Accumulated Hours:** {deletion_snapshot.get('RUN_HRS', 0):,} hrs"
                            )
                        else:
                            st.error(f"Asset {target_gcode} could not be found in the database layer.")
                            st.stop()
                    except Exception as read_err:
                        st.error(f"Failed to fetch asset metadata: {read_err}")
                        st.stop()

                st.markdown("---")

                # 3. Two-Factor Safety Confirmation Checkbox
                confirm_gate = st.checkbox(
                    f"I explicitly confirm that I want to completely delete asset **{target_gcode}** from the system database.",
                    key=f"gate_delete_{target_gcode}"
                )

                # 4. Destruction Button Execution Area
                if st.button(f"💥 PERMANENTLY DESTROY {target_gcode}", use_container_width=True, type="primary",
                             disabled=not confirm_gate):
                    try:
                        # Execute deletion row-match query against Supabase
                        supabase.table("ASSETS").delete().eq("G-CODE", target_gcode).execute()

                        # 5. Log the incident into your Tracking Architecture
                        log_audit_event(
                            target_gcode,
                            "DELETE",
                            "DECOMMISSIONED",
                            f"Asset permanently purged from system registries by administrative override.",
                            deletion_snapshot,
                            {}  # Empty object representing state after total destruction
                        )

                        st.success(f"Asset {target_gcode} successfully deleted from all network registries.")

                        # Clear system cache and trigger immediate layout sync
                        st.cache_data.clear()
                        st.rerun()

                    except Exception as delete_err:
                        st.error(f"Supabase Deletion Execution Error: {delete_err}")
            else:
                st.info("No master record data currently loaded available to purge.")
        with tab5:
            st.caption("📋 UPDATES REPORT FOR GENSET FIELD SECTIONS:")
            logs_df = get_audit_logs_df()
            if not logs_df.empty:
                logs_df['created_at'] = pd.to_datetime(logs_df['created_at']).dt.strftime('%Y-%m-%d %H:%M:%S')


                def extract_metric(row_json, key_name):
                    if isinstance(row_json, dict) and key_name in row_json:
                        val = row_json[key_name]
                        return val if val is not None and val != "----" else "—"
                    return "—"


                logs_df['From Location'] = logs_df['new_values'].apply(lambda x: extract_metric(x, 'FROM_LOCATION'))
                logs_df['To Location'] = logs_df['new_values'].apply(lambda x: extract_metric(x, 'TO_LOCATION'))
                logs_df['Reason'] = logs_df['new_values'].apply(lambda x: extract_metric(x, 'REASON'))
                logs_df['Move_date'] = logs_df['new_values'].apply(lambda x: extract_metric(x, 'MOVE_DATE'))
                logs_df['Model'] = logs_df['new_values'].apply(lambda x: extract_metric(x, 'MODEL'))
                logs_df['Serial_no'] = logs_df['new_values'].apply(lambda x: extract_metric(x, 'SERIAL_NO'))
                logs_df['KVA'] = logs_df['new_values'].apply(lambda x: extract_metric(x, 'GEN_KVA'))
                logs_df['Running Hours'] = logs_df['new_values'].apply(lambda x: extract_metric(x, 'RUN_HRS'))

                display_df = logs_df.rename(columns={
                    'created_at': 'Timestamp',
                    'Move_date': 'Move Date',
                    'Serial_no': 'Serial No',
                    'Model': 'Model',
                    'changed_by': 'Login Credentials',
                    'g_code': 'Asset ID (G-CODE)',
                    'action_type': 'Action'
                })

                target_columns = ['Move Date', 'Asset ID (G-CODE)', 'Serial No', 'Model',
                                  'KVA', 'From Location', 'To Location', 'Running Hours',
                                  'Reason', 'Login Credentials', 'Timestamp']

                # 💡 1. Convert "Move Date" column to datetimes safely for range extraction
                # Errors='coerce' turns missing/invalid dates ('—') safely into NaT (Not a Time)
                temp_move_dates = pd.to_datetime(display_df['Move Date'], errors='coerce')
                valid_move_dates = temp_move_dates.dropna()

                # Fallback boundary check if there are no valid move dates in the system yet
                import datetime

                min_date = valid_move_dates.min().date() if not valid_move_dates.empty else datetime.date.today()
                max_date = valid_move_dates.max().date() if not valid_move_dates.empty else datetime.date.today()

                # 🛠️ Layout columns for filters
                col1, col2 = st.columns([1, 1])

                with col1:
                    # 💡 2. Date Range Picker mapped to Move Date
                    date_range = st.date_input(
                        "📅 Filter by Move Date Range:",
                        value=(min_date, max_date),
                        min_value=min_date,
                        max_value=max_date,
                        key="tab5_move_date_range"
                    )

                with col2:
                    # 💡 3. Dynamic Unique Options for Multi-Select Filter
                    searchable_columns = ['Asset ID (G-CODE)', 'From Location', 'To Location', 'Model', 'Reason']
                    unique_options = set()
                    for col in searchable_columns:
                        if col in display_df.columns:
                            unique_options.update(display_df[col].dropna().astype(str).unique())
                    unique_options = sorted([opt for opt in unique_options if opt not in ["—", "----", ""]])

                    selected_queries = st.multiselect(
                        "🔍 Multi-Select Global Filter:",
                        options=unique_options,
                        placeholder="Choose terms...",
                        key="tab5_multi_search"
                    )

                    # 🏃‍♂️ 4. Execution of Filters
                    filtered_df = display_df.copy()

                    # Step A: Apply Move Date Filter
                    if isinstance(date_range, (tuple, list)) and len(date_range) == 2:
                        start_date, end_date = date_range

                        # Map back to our temporary datetime series to cleanly filter matching boundaries
                        filtered_df = filtered_df[
                            (temp_move_dates.dt.date >= start_date) &
                            (temp_move_dates.dt.date <= end_date)
                            ]

                    # Step B: Apply Multi-Select Global Search Filter
                    if selected_queries:
                        selected_queries_lower = [str(q).lower() for q in selected_queries]
                        mask = filtered_df[target_columns].astype(str).apply(
                            lambda row: row.str.lower().apply(
                                lambda cell: any(q in cell for q in selected_queries_lower)).any(),
                            axis=1
                        )
                        filtered_df = filtered_df[mask]

                    # Render Final Output Table
                if not filtered_df.empty:
                    designed_logs_df = filtered_df[target_columns].style.apply(style_zebra_rows, axis=None)
                    st.dataframe(designed_logs_df, use_container_width=True, hide_index=True)
                else:
                    st.warning("No records matched your combined move date range and search criteria.")

            else:
                st.info("No audit transactions logged inside tracking structures yet.")

        with tab6:
            st.caption(
                "Real-time monitoring, diagnostic breakdowns, and location-based asset tracking for field operations.")

            try:
                fleet_df = get_assets_df()
            except Exception as e:
                st.error(f"Error fetching workshop data streams: {e}")
                fleet_df = pd.DataFrame()

            if not fleet_df.empty:
                if 'GEN_KVA' in fleet_df.columns:
                    fleet_df['GEN_KVA'] = pd.to_numeric(fleet_df['GEN_KVA'], errors='coerce').fillna(0).astype(int)
                if 'RUN_HRS' in fleet_df.columns:
                    fleet_df['RUN_HRS'] = pd.to_numeric(fleet_df['RUN_HRS'], errors='coerce').fillna(0).astype(int)


                def safe_str(val):
                    if val is None or pd.isna(val) or str(val).strip().upper() == "NONE":
                        return "-"
                    return str(val).strip().encode('ascii', 'ignore').decode('ascii')


                # Dynamic toggle configuration
                search_mode = st.radio(
                    "Choose Search Vector:",
                    ["🔍 Search by G-CODE", "📍 Search by Location Pipeline"],
                    horizontal=True
                )

                # =========================================================================
                # MODE A: SEARCH BY SPECIFIC G-CODE
                # =========================================================================
                if search_mode == "🔍 Search by G-CODE":
                    default_choice = "--- SELECT A SPECIFIC G-CODE FOR FULL PROFILE ---"
                    all_gcodes = [default_choice] + sorted(fleet_df['G-CODE'].dropna().unique().tolist())

                    if "fm_gcode_focus_dropdown" not in st.session_state:
                        st.session_state["fm_gcode_focus_dropdown"] = default_choice

                    selected_gcode_focus = st.selectbox(
                        "SELECT G-CODE :",
                        options=all_gcodes,
                        key="fm_gcode_focus_dropdown"
                    )

                    if selected_gcode_focus != default_choice:
                        isolated_asset_df = fleet_df[fleet_df['G-CODE'] == selected_gcode_focus]

                        if not isolated_asset_df.empty:
                            asset_row = isolated_asset_df.iloc[0]

                            st.caption(f" 📋 Full Profile File: {selected_gcode_focus}")
                            prof_stage = safe_str(asset_row.get('USER'))
                            prof_fault = safe_str(asset_row.get('REASON'))

                            status_color = "green" if "READY" in prof_stage.upper() or "COMPLETED" in prof_stage.upper() else (
                                "orange" if "QC" in prof_stage.upper() else "blue")
                            st.markdown(
                                f"**Current Status:** :{status_color}[**{prof_stage}**] | **Primary Diagnostics:** *{prof_fault}*")

                            progress_val = 1.0 if "READY" in prof_stage.upper() or "COMPLETED" in prof_stage.upper() else (
                                0.7 if "QC" in prof_stage.upper() else (0.4 if "REPAIR" in prof_stage.upper() else 0.1))
                            st.progress(progress_val)

                            st.caption("#### 🔍 Master Specifications Card")
                            card_col1, card_col2, card_col3 = st.columns(3)

                            with card_col1:
                                st.info("⚙️ Mechanical Inventory Spec")
                                st.markdown(f"**Asset Code (G-CODE):** `{safe_str(asset_row.get('G-CODE'))}`")
                                st.markdown(f"**GEN_TYPE:** {safe_str(asset_row.get('TYPE'))}")
                                st.markdown(f"**GEN_Model:** {safe_str(asset_row.get('MODEL'))}")
                                st.markdown(f"**Capacity Rating:** {int(asset_row.get('GEN_KVA', 0)):,} KVA")
                                st.markdown(f"**Operating Timeline:** {int(asset_row.get('RUN_HRS', 0)):,} Hours")

                            with card_col2:
                                st.warning("🔧 Workshop Diagnostics")
                                st.markdown(f"**Assigned Stage:** `{prof_stage}`")
                                st.markdown(f"**FIELD_LOCATION:** {safe_str(asset_row.get('FIELD'))}")
                                st.markdown(f"**Reported Mechanical Issue:** *{prof_fault}*")
                                p_notes = asset_row.get('PURPOSE', asset_row.get('REMARKS', '-'))
                                st.markdown(f"**Technician Maintenance Action Logs:** {safe_str(p_notes)}")

                            with card_col3:
                                st.success("🚚 Dispatch & Logistics Targets")
                                st.markdown(f"**Target Site Post-Repair:** {safe_str(asset_row.get('TO_LOCATION'))}")
                                if 'MOVE_DATE' in asset_row:
                                    st.markdown(
                                        f"**Last Movement Log Timestamp:** {safe_str(asset_row.get('MOVE_DATE'))}")
                                if 'SERIAL_NO' in asset_row or 'SERIAL' in asset_row:
                                    s_val = asset_row.get('SERIAL_NO', asset_row.get('SERIAL', '-'))
                                    st.markdown(f"**Chassis / Engine Serial No:** `{safe_str(s_val)}`")

                            if st.button("⬅️ Back to Global Workshop Fleet Overview Table", use_container_width=True):
                                st.session_state["fm_gcode_focus_dropdown"] = default_choice
                                st.rerun()

                # =========================================================================
                # MODE B: SEARCH BY LOCATION PIPELINE (e.g., BG-0002)
                # =========================================================================
                elif search_mode == "📍 Search by Location Pipeline":
                    # Consolidate unique location tags from both tracking vectors for clean reference lookup
                    raw_fields = fleet_df['FIELD'].dropna().unique().tolist() + fleet_df[
                        'TO_LOCATION'].dropna().unique().tolist()
                    unique_fields = sorted(
                        list(set([str(x).strip() for x in raw_fields if str(x).strip() not in ["", "----", "-"]])))

                    default_loc_choice = "--- SELECT OR TYPE A LOCATION ---"
                    loc_options = [default_loc_choice] + unique_fields

                    selected_loc_box = st.selectbox(
                        "Quick Select Active Location:",
                        options=loc_options,
                        help="Select an existing field location from the active registry."
                    )

                    text_target_location = st.text_input(
                        "Or Type Custom Location / Code Manually:",
                        placeholder="e.g. BG-0002, Field Alpha...",
                        help="Type any part of the field location name or site ID code."
                    ).strip()

                    target_location = text_target_location if text_target_location else (
                        "" if selected_loc_box == default_loc_choice else selected_loc_box
                    )

                    if target_location:
                        # Multi-column mapping logic check across fields or target coordinates
                        filtered_location_df = fleet_df[
                            fleet_df['FIELD'].astype(str).str.contains(target_location, case=False, na=False) |
                            fleet_df['TO_LOCATION'].astype(str).str.contains(target_location, case=False, na=False)
                            ]

                        if not filtered_location_df.empty:
                            total_found = len(filtered_location_df)
                            st.success(
                                f"📍 Found **{total_found}** asset(s) currently posted at location: `{target_location}`")

                            for idx, row in filtered_location_df.iterrows():
                                gcode_lbl = safe_str(row.get('G-CODE'))
                                stage_lbl = safe_str(row.get('USER'))

                                with st.expander(f"📦 Asset ID: {gcode_lbl} — Current Stage: {stage_lbl}",
                                                 expanded=True):
                                    loc_col1, loc_col2, loc_col3 = st.columns(3)

                                    with loc_col1:
                                        st.info("⚙️ Mechanical Spec")
                                        st.markdown(f"**G-CODE:** `{gcode_lbl}`")
                                        st.markdown(
                                            f"**Model:** {safe_str(row.get('TYPE'))} / {safe_str(row.get('MODEL'))}")
                                        st.markdown(f"**Rating:** {int(row.get('GEN_KVA', 0)):,} KVA")

                                    with loc_col2:
                                        st.warning("🔧 Operational Diagnostics")
                                        st.markdown(f"**Registered Field:** `{safe_str(row.get('FIELD'))}`")
                                        st.markdown(f"**Current Run Hours:** {int(row.get('RUN_HRS', 0)):,} Hrs")
                                        st.markdown(f"**Reported Issue:** *{safe_str(row.get('REASON'))}*")

                                    with loc_col3:
                                        st.success("🚚 Dispatch Targets")
                                        st.markdown(f"**Target Destination:** {safe_str(row.get('TO_LOCATION'))}")
                                        p_notes_loc = row.get('PURPOSE', row.get('REMARKS', '-'))
                                        st.markdown(f"**Technician Logs:** {safe_str(p_notes_loc)}")
                        else:
                            st.warning(f"No active asset matches found for location criteria: `{target_location}`")

    # =====================================================================
    # 3. SETTINGS MATRIX PLATFORM
    # =====================================================================
    elif navigation_target == "SETTINGS":
        st.title("⚙️ System Operational Settings Configuration")
        cfg_tab1, cfg_tab2, cfg_tab3, cfg_tab4 = st.tabs(
            ["Active Models Mapping", "Upsert Model Profiles", "🛠️ MANAGE FIELD LOGISTICS REGISTRIES", "USER UPDATE"])

        with cfg_tab1:
            st.caption("Active Mapping Rules Matrix (Loaded from Database)")
            if not mappings_df.empty:
                st.dataframe(mappings_df[['type', 'model', 'kva']], use_container_width=True, hide_index=True)

        with cfg_tab2:
            st.subheader("Manage Specifications Matrix Rule")
            with st.form("EDIT_MAPPING_MATRIX_FORM", clear_on_submit=True):
                m_col1, m_col2, m_col3 = st.columns(3)
                with m_col1:
                    cfg_type = st.text_input("Manufacturer Type Name (e.g., PERKINS):").strip().upper()
                with m_col2:
                    cfg_model = st.text_input("Engine Model Code (e.g., 2506A):").strip().upper()
                with m_col3:
                    cfg_kva = st.number_input("Standard Default KVA Rating:", min_value=0, value=100, step=10)

                if st.form_submit_button("UPSERT CONFIGURATION PROFILE"):
                    if not cfg_type or not cfg_model:
                        st.error("Both Parameters are required.")
                    else:
                        try:
                            supabase.table("asset_mappings").upsert(
                                {"type": cfg_type, "model": cfg_model, "kva": cfg_kva},
                                on_conflict="type,model"
                            ).execute()
                            st.success("Matrix rule updated successfully!")
                            st.cache_data.clear()
                            st.rerun()
                        except Exception as e:
                            st.error(f"Failed to apply matrix update: {e}")

        with cfg_tab3:
            st.subheader("Dynamic Field/Area/Location Cascading Matrix Registry")
            st.caption("Add or delete options mapping paths to configure dropdown availability parameters seamlessly.")

            st.markdown("##### Active Logistical Configuration Routing Path Ledger:")
            if not routing_df.empty:
                display_routing = routing_df.copy().rename(
                    columns={'field_name': 'FIELD Option', 'area_name': 'AREA Option',
                             'location_name': 'TO_LOCATION Option'})
                st.dataframe(display_routing[['id', 'FIELD Option', 'AREA Option', 'TO_LOCATION Option']],
                             use_container_width=True, hide_index=True)
            else:
                st.info("No cascading routing rules generated inside database structures yet.")

            st.markdown("---")
            act_col1, act_col2 = st.columns(2)

            with act_col1:
                st.markdown("➕ **Inject New Cascading Route Validation Options**")
                with st.form("ADD_NEW_FIELD_ROUTE_FORM", clear_on_submit=True):
                    in_field = st.text_input("Target FIELD Identifier Name:",
                                             placeholder="e.g., SABRIYA_YARD").strip().upper()
                    in_area = st.text_input("Target AREA Identifier Code:", placeholder="e.g., SBY").strip().upper()
                    in_loc = st.text_input("Target TO_LOCATION Identifier Node:",
                                           placeholder="e.g., MN-0025").strip().upper()

                    if st.form_submit_button("REGISTER ROUTE PATH OPTION", use_container_width=True):
                        if not in_field or not in_area or not in_loc:
                            st.error("Validation Halt: All path node identities must be completely specified.")
                        else:
                            try:
                                supabase.table("field_routing_matrix").insert(
                                    {"field_name": in_field, "area_name": in_area, "location_name": in_loc}).execute()
                                st.success(
                                    f"Successfully pinned structural logistics pathway path entry node layout mapping!")
                                st.cache_data.clear()
                                st.rerun()
                            except Exception as ex:
                                st.error(f"Failed to record logistical pathway matrix configuration payload: {ex}")

            with act_col2:
                st.markdown("🗑️ **Prune Existing Matrix Route Profile**")
                if not routing_df.empty:
                    routing_df['label_str'] = "FIELD: " + routing_df['field_name'] + " ➡️ AREA: " + routing_df[
                        'area_name'] + " ➡️ LOC: " + routing_df['location_name']
                    id_map = dict(zip(routing_df['label_str'], routing_df['id']))

                    selected_route_str = st.selectbox("Select Target Path Sequence Configuration to Purge:",
                                                      options=sorted(id_map.keys()), key="route_purge_selectbox")
                    target_route_id = id_map[selected_route_str]

                    if st.button("🔥 DISMANTLE SPECIFIED LOGISTICAL PATHWAY", use_container_width=True):
                        try:
                            supabase.table("field_routing_matrix").delete().eq("id", int(target_route_id)).execute()
                            st.success(
                                "Logistical pathway component successfully dropped out of database configuration profiles!")
                            st.cache_data.clear()
                            st.rerun()
                        except Exception as ex:
                            st.error(f"Purge Execution Denied: {ex}")
                else:
                    st.info("No routing matrices available to drop.")

        with cfg_tab4:
            with st.expander("USER SETTINGS", expanded=True):
                st.subheader("⚙️ User Assignment Registry Control")
                st.caption("Add or decommission asset operators/users without modifying application source code.")

                reg_col1, reg_col2 = st.columns(2)

                with reg_col1:
                    st.markdown("##### Add New Registry Entry")
                    new_reg_user = st.text_input("Operator / User Name:", key="new_user_reg_input").strip().upper()
                    if st.button("➕ Register User", use_container_width=True):
                        if not new_reg_user or new_reg_user == '----':
                            st.warning("Please type a valid user identification string.")
                        else:
                            try:
                                supabase.table("USER_REGISTRY").insert({'user_name': new_reg_user}).execute()
                                st.success(f"'{new_reg_user}' integrated successfully!")
                                st.cache_data.clear()
                                st.rerun()
                            except Exception as err:
                                st.error(f"Failed to write record: {err}")

                with reg_col2:
                    st.markdown("##### Delete Registry Entry")
                    # Force dynamic evaluation from dynamic function context
                    current_live_users = load_user_registry()
                    deletion_options = [u for u in current_live_users if u != '----']
                    user_to_delete = st.selectbox("Select User to Remove:", options=deletion_options,
                                                  key="delete_user_reg_select")

                    if st.button("❌ Drop User Profile", use_container_width=True):
                        if user_to_delete:
                            try:
                                supabase.table("USER_REGISTRY").delete().eq('user_name', user_to_delete).execute()
                                st.success(f"'{user_to_delete}' purged from central systems.")
                                st.cache_data.clear()
                                st.rerun()
                            except Exception as err:
                                st.error(f"Failed to drop database link: {err}")

    # =====================================================================
    # 4. SYSTEM ADMINISTRATIVE AUDIT MONITOR LOGS
    # =====================================================================
    elif navigation_target == "AUDIT LOGS":
        st.title("⚙️ System Administrative Control Settings")
        logs_df = get_audit_logs_df()

        if not logs_df.empty:
            logs_df['created_at'] = pd.to_datetime(logs_df['created_at']).dt.strftime('%Y-%m-%d %H:%M:%S')


            def extract_metric(row_json, key_name):
                if isinstance(row_json, dict) and key_name in row_json:
                    val = row_json[key_name]
                    return val if val is not None and val != "----" else "—"
                return "—"


            logs_df['From Location'] = logs_df['new_values'].apply(lambda x: extract_metric(x, 'FROM_LOCATION'))
            logs_df['To Location'] = logs_df['new_values'].apply(lambda x: extract_metric(x, 'TO_LOCATION'))
            logs_df['Reason'] = logs_df['new_values'].apply(lambda x: extract_metric(x, 'REASON'))
            logs_df['KVA Rating'] = logs_df['new_values'].apply(lambda x: extract_metric(x, 'GEN_KVA'))
            logs_df['Running Hours'] = logs_df['new_values'].apply(lambda x: extract_metric(x, 'RUN_HRS'))

            search_gcode = st.text_input("🔍 Quick-Filter Ledger by Asset ID (G-CODE):", placeholder="e.g., G-009",
                                         key="standalone_audit_ledger_search").strip().upper()
            filtered_df = logs_df[
                logs_df['g_code'].str.upper().str.contains(search_gcode, na=False)] if search_gcode else logs_df.copy()

            display_df = filtered_df.rename(
                columns={'created_at': 'Timestamp', 'changed_by': 'Login Credentials', 'g_code': 'Asset ID (G-CODE)',
                         'action_type': 'Action'})
            target_columns = ['Timestamp', 'Asset ID (G-CODE)', 'From Location', 'To Location', 'Running Hours',
                              'Reason', 'KVA Rating', 'Login Credentials']

            designed_logs_df = display_df[target_columns].style.apply(style_zebra_rows, axis=None)
            st.dataframe(designed_logs_df, use_container_width=True, hide_index=True)
        else:
            st.info("No audit transactions logged inside tracking structures yet.")

        st.markdown("---")
        st.error("⚠️ CRITICAL ADMINISTRATIVE ACTIONS: PURGE ENGINE AUDIT LEDGER")

        purge_mode = st.radio(
            "Select Purge Strategy:",
            options=["Retain Recent Logs (By Date)", "💥 COMPLETE SYSTEM WIPE (Delete Everything Automatically)"],
            index=0,
            key="admin_purge_strategy_choice_standalone"
        )

        purge_all_flag = False
        retention_days = 30

        if "Retain Recent Logs" in purge_mode:
            retention_days = st.number_input(
                "Select Log Retention Window (Days to Keep):",
                min_value=1, max_value=365, value=30, step=1,
                key="admin_retention_days_input_standalone"
            )
        else:
            purge_all_flag = True
            st.warning(
                "‼️ WARNING: Choosing this option will erase every row in the AUDIT_LOGS table. This action is permanent.")

        confirm_purge = st.checkbox("I explicitly authorize the irreversible destruction of these data registries.",
                                    value=False, key="ui_lock_purge_gate_standalone")

        if st.button("🔥 EXECUTE SYSTEM PURGE MANDATE", use_container_width=True, disabled=not confirm_purge,
                     key="execute_purge_btn_standalone"):
            with st.spinner("Wiping database tracking logs..."):
                purge_result = delete_old_audit_logs(days_retention=retention_days, purge_all=purge_all_flag)
                if purge_result["status"] == "success":
                    st.cache_data.clear()
                    st.success(f"Successfully cleared out {purge_result['row_count']} logs from the database!")
                    time.sleep(1)
                    st.rerun()
                else:
                    st.error(f"Purge process failure: {purge_result.get('message')}")

        # =====================================================================
        # 5. UNIMPLEMENTED PLACEHOLDER ROUTING TARGETS
        # =====================================================================
    #=====================================================================
    #workshop
    #=====================================================================
    elif navigation_target == "WORKSHOP":
        df = get_assets_df()
        if not df.empty:
            try:
                import pandas as pd
                import datetime

                # --- CRITICAL: SANITIZE DATA TYPES TO PREVENT RUNTIME ERRORS ---
                if 'GEN_KVA' in df.columns:
                    df['GEN_KVA'] = pd.to_numeric(df['GEN_KVA'], errors='coerce').fillna(0).astype(int)
                if 'RUN_HRS' in df.columns:
                    df['RUN_HRS'] = pd.to_numeric(df['RUN_HRS'], errors='coerce').fillna(0)

                # --- ISOLATED WORKSHOP FILTER LAYER ---
                if 'USER' in df.columns and 'TO_LOCATION' in df.columns:
                    workshop_df = df[(df['USER'] == 'WORKSHOP') & (df['TO_LOCATION'] == 'WORKSHOP')].copy()
                else:
                    workshop_df = pd.DataFrame()

                # --- 💡 NEW: CALCULATE DURATION METRIC ("Days in Workshop") ---
                if not workshop_df.empty and 'MOVE_DATE' in workshop_df.columns:
                    # Safely parse the move date column; invalid formats/dashes turn into NaT
                    parsed_move_dates = pd.to_datetime(workshop_df['MOVE_DATE'], errors='coerce')

                    # Calculate days between today's date and the move date
                    today_dt = pd.Timestamp(datetime.date.today())

                    # .dt.days extracts just the integer count of days
                    workshop_df['Days in Workshop'] = (today_dt - parsed_move_dates).dt.days

                    # Clean up any missing dates or future clock errors cleanly
                    workshop_df['Days in Workshop'] = workshop_df['Days in Workshop'].apply(
                        lambda x: f"{int(x)} Days" if pd.notna(x) and x >= 0 else "—"
                    )
                else:
                    workshop_df['Days in Workshop'] = "—"

                # --- RE-CALCULATE METRICS SPECIFIC TO WORKSHOP DATA ---
                v_TT = len(workshop_df)
                total_kva = workshop_df['GEN_KVA'].sum() if 'GEN_KVA' in workshop_df.columns else 0
                avg_runtime = workshop_df['RUN_HRS'].mean() if 'RUN_HRS' in workshop_df.columns else 0

                # --- POWER BI FEATURE: EXECUTIVE SUMMARY METRICS FIRST ---
                st.subheader("🛠️ Workshop Status Control Dashboard")
                st.caption("📋 Real-time metrics for assets physically stationed inside the main workshop")

                kpi1, kpi2, kpi3 = st.columns(3)
                with kpi1:
                    st.metric(label="📟 UNITS IN WORKSHOP", value=f"{v_TT:,} Units", border=True)
                with kpi2:
                    st.metric(label="⚡ TOTAL UNDER-REPAIR CAPACITY", value=f"{total_kva:,} KVA", border=True)
                with kpi3:
                    st.metric(label="⏳ AVG ACCUMULATED RUN HOURS",
                              value=f"{int(avg_runtime):,} Hrs" if pd.notna(avg_runtime) else "0 Hrs", border=True)

                st.markdown("---")

                # --- LIVE VIEWPORT DATAFRAME ---
                st.markdown("##### 📋 Workshop Assets Ledger Viewport")
                if not workshop_df.empty:
                    # Rename columns dynamically for structured enterprise presentations
                    display_workshop_df = workshop_df.rename(columns={
                        'created_at': 'Timestamp',
                        'MOVE_DATE': 'Move Date',
                        'SERIAL_NO': 'Serial No',
                        'MODEL': 'Model',
                        'TYPE': 'Manufacturer TYPE',
                        'G-CODE': 'Asset ID (G-CODE)',
                        'GEN_KVA': 'Rating (KVA)',
                        'RUN_HRS': 'Running Hours',
                        'REASON': 'Repair Remarks / Reason'
                    })

                    # 💡 Added 'Days in Workshop' right next to the Move Date for optimal scannability
                    target_columns = [
                        'Asset ID (G-CODE)', 'Manufacturer TYPE', 'Model', 'Rating (KVA)',
                        'Running Hours', 'Move Date', 'Days in Workshop', 'Repair Remarks / Reason', 'Timestamp'
                    ]

                    # Check for existing column compatibility slice safely
                    available_cols = [c for c in target_columns if c in display_workshop_df.columns]

                    # Apply styling and render layout frame
                    designed_workshop_df = display_workshop_df[available_cols].style.apply(style_zebra_rows, axis=None)
                    st.dataframe(designed_workshop_df, use_container_width=True, hide_index=True)
                else:
                    st.info("Excellent! There are currently zero assets flagged under Workshop repair constraints.")

            except Exception as e:
                st.error(f"Error compiling Workshop tracking registry views: {e}")
        else:
            st.info("No equipment inventory assets found inside database registries.")
    #===========================================================================================
    #STORES AND PARTS
    #===========================================================================================
    elif navigation_target == "STORES & PARTS":
        st.subheader("📦 Spare Parts Store Inventory & Dispatch Control")

        # --- 1. DYNAMIC METRICS AGGREGATION ENGINE ---
        try:
            # Fetch current store inventory profiles
            parts_query = supabase.table("SPARE_PARTS").select("*").execute()
            parts_list = parts_query.data if parts_query.data else []

            # Fetch permanent historical tracking logs for the parts store
            audit_query = supabase.table("AUDIT_LOGS").select("*").in_("action_type",
                                                                       ["STORE_INBOUND", "STORE_OUTBOUND"]).execute()
            audit_list = audit_query.data if audit_query.data else []
        except Exception as fetch_err:
            st.error(f"Failed to synchronize parts inventory streams: {fetch_err}")
            parts_list = []
            audit_list = []

        # --- INTAKE SEGMENT CALCULATIONS ---
        total_line_items = len(parts_list)
        total_physical_pieces = sum([int(item.get('quantity', 0)) for item in parts_list])

        # --- DISPATCH SEGMENT CALCULATIONS ---
        dispatch_logs_raw = [log for log in audit_list if log.get('action_type') == 'STORE_OUTBOUND']
        total_dispatched_transactions = len(dispatch_logs_raw)

        # Safely parse structural strings or snapshots to count total hardware pieces sent out to operations
        import json

        total_pieces_dispatched = 0
        for log in dispatch_logs_raw:
            try:
                # Look inside the previous snapshot vs new snapshot to deduce volume moved
                old_snap = log.get('old_snapshot', {})
                new_snap = log.get('new_snapshot', {})
                if isinstance(old_snap, str): old_snap = json.loads(old_snap)
                if isinstance(new_snap, str): new_snap = json.loads(new_snap)

                diff = int(old_snap.get('quantity', 0)) - int(new_snap.get('quantity', 0))
                if diff > 0:
                    total_pieces_dispatched += diff
            except:
                pass

        # --- 2. CLEANLY SEPARATED METRICS GROUPS ---

        # GROUP A: INBOUND INVENTORY CAPACITY METRICS
        st.markdown("##### 📥 Current Storehouse Inventory Totals")
        in_kpi1, in_kpi2 = st.columns(2)
        with in_kpi1:
            st.metric(label="🗂️ UNIQUE CATALOGED PART NUMBERS", value=f"{total_line_items:,} Items", border=True)
        with in_kpi2:
            st.metric(label="📦 TOTAL PHYSICAL PIECES ON SHELVES", value=f"{total_physical_pieces:,} Units", border=True)

        st.markdown("---")

        # GROUP B: OUTBOUND LEAN FLEET PERFORMANCE METRICS
        st.markdown("##### 📤 Cumulative Historical Fleet Dispatch Totals")
        out_kpi1, out_kpi2 = st.columns(2)
        with out_kpi1:
            st.metric(label="🚀 ACTIVE DISPATCH TRANSACTIONS RUN", value=f"{total_dispatched_transactions:,} Orders",
                      border=True)
        with out_kpi2:
            st.metric(label="🔧 TOTAL SPARE PIECES DISTRIBUTED TO FLEET",
                      value=f"{total_pieces_dispatched:,} Parts Issued", border=True)

        st.markdown("---")

        # --- 3. DUAL-STREAM TRANSACTION FORMS ---
        action_tab1, action_tab2 = st.tabs(["📥 Receive / Add Stock", "📤 Dispatch / Issue Stock"])

        # TRACK 1: INVENTORY INTAKE FORM
        with action_tab1:
            st.markdown("##### Log New Spare Parts Into Storage")
            with st.form(key="add_spare_parts_form"):
                in_part_no = st.text_input("Part Number (Unique Identifier) *:", key="st_in_part_no").strip().upper()
                in_name = st.text_input("Name of Spare Part *:", key="st_in_name").strip()
                in_qty = st.number_input("Quantity Received *:", min_value=1, step=1, value=1, key="st_in_qty")
                in_loc = st.text_input("Storage Location Shelf / Bin *:", key="st_in_loc").strip()

                submit_add = st.form_submit_button("ADD STOCK TO INVENTORY", use_container_width=True)

                if submit_add:
                    if not in_part_no or not in_name or not in_loc:
                        st.error("❌ Action Blocked: Please complete all required configuration fields (*).")
                    else:
                        try:
                            exist_q = supabase.table("SPARE_PARTS").select("*").eq("part_number", in_part_no).execute()
                            if exist_q.data:
                                old_record = exist_q.data[0]
                                new_qty = int(old_record.get('quantity', 0)) + int(in_qty)
                                supabase.table("SPARE_PARTS").update({"quantity": new_qty}).eq("part_number",
                                                                                               in_part_no).execute()
                                new_record = {**old_record, "quantity": new_qty}

                                log_audit_event(
                                    in_part_no, "STORE_INBOUND", "STOCK_INCREMENT",
                                    f"Added {in_qty} units of existing stock. Total inventory: {new_qty}.",
                                    old_record, new_record
                                )
                            else:
                                new_payload = {"part_number": in_part_no, "spare_name": in_name,
                                               "quantity": int(in_qty), "location": in_loc}
                                supabase.table("SPARE_PARTS").insert(new_payload).execute()

                                log_audit_event(
                                    in_part_no, "STORE_INBOUND", "NEW_ITEM",
                                    f"Registered brand new spare part asset: {in_qty} units.",
                                    {}, new_payload
                                )
                            st.success(
                                f"📦 Inventory matrix verified! {in_qty} units of Part [{in_part_no}] synchronized.")
                            st.cache_data.clear()
                            st.rerun()
                        except Exception as err:
                            st.error(f"Inventory DB execution fault: {err}")

        # TRACK 2: DISTRIBUTION / DISPATCH FORM
        with action_tab2:
            st.markdown("##### Issue Spares to Field Assignments / Workorders")
            if parts_list:
                parts_options = {
                    item['part_number']: f"{item['part_number']} — {item['spare_name']} (Available: {item['quantity']})"
                    for item in parts_list}

                with st.form(key="dispatch_parts_form"):
                    selected_part_no = st.selectbox(
                        "Select Spare Part to Dispatch:",
                        options=list(parts_options.keys()),
                        format_func=lambda x: parts_options[x],
                        key="st_out_part_no"
                    )
                    out_qty = st.number_input("Quantity to Issue:", min_value=1, step=1, value=1, key="st_out_qty")
                    out_target = st.text_input("Destination Place / Target Asset ID (e.g. G-CODE) *:",
                                               key="st_out_target").strip().upper()
                    out_remarks = st.text_area("Dispatch Purpose / Job Card Remarks:", key="st_out_remarks").strip()

                    submit_dispatch = st.form_submit_button("AUTHORIZE INVENTORY DISPATCH", use_container_width=True)

                    if submit_dispatch:
                        if not out_target:
                            st.error("❌ Action Blocked: You must identify the target machine or destination asset ID.")
                        else:
                            current_item = next(item for item in parts_list if item['part_number'] == selected_part_no)
                            available_stock = int(current_item.get('quantity', 0))

                            if out_qty > available_stock:
                                st.error(f"❌ Shortage Error: Requested: {out_qty}, Available: {available_stock}")
                            else:
                                try:
                                    final_qty = available_stock - int(out_qty)
                                    supabase.table("SPARE_PARTS").update({"quantity": final_qty}).eq("part_number",
                                                                                                     selected_part_no).execute()
                                    updated_item = {**current_item, "quantity": final_qty}

                                    log_audit_event(
                                        selected_part_no,
                                        "STORE_OUTBOUND",
                                        "STOCK_DEDUCTION",
                                        f"Issued {out_qty} units out to Asset/Place: {out_target}. Purpose: {out_remarks}",
                                        current_item,
                                        updated_item
                                    )
                                    st.success(
                                        f"🚀 Dispatch secure! {out_qty} units mapped against reference target: {out_target}.")
                                    st.cache_data.clear()
                                    st.rerun()
                                except Exception as out_err:
                                    st.error(f"Dispatch update failure: {out_err}")
            else:
                st.info("No stock records currently exist inside storage shelves to display for distribution cycles.")

        st.markdown("---")

        # --- 4. LIVE LOG VIEWPORT (THE PERMANENT DISPATCH HISTORICAL RECORD) ---
        st.markdown("##### 📋 Permanent Storage Dispatch & Transaction History")
        if audit_list:
            import pandas as pd
            import json
            import re

            audit_df = pd.DataFrame(audit_list)

            # Filter down specifically to outbound dispatches
            if not audit_df.empty and 'action_type' in audit_df.columns:
                dispatch_logs = audit_df[audit_df['action_type'] == 'STORE_OUTBOUND'].copy()
            else:
                dispatch_logs = pd.DataFrame()

            if not dispatch_logs.empty:
                # Sort with the latest transaction appearing at the top
                if 'created_at' in dispatch_logs.columns:
                    dispatch_logs['created_at'] = pd.to_datetime(dispatch_logs['created_at'])
                    dispatch_logs = dispatch_logs.sort_values(by='created_at', ascending=False)
                    dispatch_logs['created_at'] = dispatch_logs['created_at'].dt.strftime('%Y-%m-%d %H:%M')


                # 🛠️ PARSE PART NUMBER DYNAMICALLY
                def parse_part_number(row):
                    # Check explicit columns
                    for key in ['asset_gcode', 'gcode', 'part_number']:
                        if key in row and pd.notna(row[key]) and str(row[key]).strip() not in ["", "—"]:
                            return str(row[key])

                    # Check text/notes descriptions safely
                    for key in ['notes', 'remarks', 'reason', 'details']:
                        if key in row and pd.notna(row[key]):
                            notes_str = str(row[key])
                            match = re.search(r"Part\s*\[([^\]]+)\]", notes_str, re.IGNORECASE)
                            if match:
                                return match.group(1)
                    return "—"


                # 🛠️ EXTRACT REMAINING STOCK BALANCE VIA LIVE INVENTORY OR SNAPSHOTS
                def parse_balance(row, parts_list):
                    try:
                        after_state = row.get('new_snapshot', {})
                        if isinstance(after_state, str):
                            after_state = json.loads(after_state)
                        if after_state:
                            for q_key in ['quantity', 'qty', 'Stock', 'balance']:
                                if q_key in after_state and after_state[q_key] != 0:
                                    return f"{int(after_state[q_key])} Pcs"
                    except:
                        pass

                    # Cross-reference with our live pulled parts list
                    p_num = parse_part_number(row)
                    if p_num != "—":
                        live_match = next((item for item in parts_list if str(item.get('part_number')) == p_num), None)
                        if live_match and 'quantity' in live_match:
                            return f"{int(live_match['quantity'])} Pcs"

                    return "—"


                # 🛠️ SAFE PARSING FOR NOTES / DETAILS COLUMN
                def parse_destination(row):
                    for key in ['notes', 'remarks', 'reason', 'details', 'description']:
                        if key in row and pd.notna(row[key]) and str(row[key]).strip() != "":
                            return str(row[key])
                    return "—"


                # Apply processing calculations safely using row loops (.get/apply) to prevent crashes
                dispatch_logs['Part Number'] = dispatch_logs.apply(parse_part_number, axis=1)
                dispatch_logs['Stock Balance'] = dispatch_logs.apply(lambda r: parse_balance(r, parts_list), axis=1)
                dispatch_logs['Dispatch Details & Destination'] = dispatch_logs.apply(parse_destination, axis=1)

                dispatch_logs['Issued By'] = dispatch_logs.apply(
                    lambda r: str(r.get('operator_user', r.get('user_id', 'SYSTEM'))), axis=1
                )

                # Map Timestamp column safely
                if 'created_at' in dispatch_logs.columns:
                    dispatch_logs['Timestamp'] = dispatch_logs['created_at']
                else:
                    dispatch_logs['Timestamp'] = "—"

                # Target exact final display columns
                target_cols = [
                    'Timestamp',
                    'Part Number',
                    'Stock Balance',
                    'Dispatch Details & Destination',
                    'Issued By'
                ]

                # Final data filter alignment pass
                available_cols = [c for c in target_cols if c in dispatch_logs.columns]
                final_display_df = dispatch_logs[available_cols]

                styled_logs = final_display_df.style.apply(style_zebra_rows, axis=None)
                st.dataframe(styled_logs, use_container_width=True, hide_index=True)
            else:
                st.info("No historical dispatch operations logged yet.")
        else:
            st.info("No transaction logs registered under warehouse inbound or outbound records.")


    #======================================================================================================
    elif navigation_target == "MAINTENANCE":
        st.info("Maintenance schedules workflow manager pending engineering configuration.")
    #=============================================================================================
    elif navigation_target == "FLEET MANAGEMENT":
        st.info("Fleet management pending engineering configuration.")

    elif navigation_target == "REMOTE TELEMETRY":
        st.info("Scada/IoT streaming pipeline processing endpoints interface metrics offline.")
    #=======================================================================================================
    elif navigation_target == "REPORTS":
        st.subheader("📋 Centralized Field Operations Reporting Cockpit")
        st.caption("Generate, filter, and export live executive-ready reports across fleet segments")

        # --- 1. COLLECT DATAFRAME SOURCES ---
        try:
            fleet_df = get_assets_df()

            # Pull audit logs safely for transactional metrics
            audit_query = supabase.table("AUDIT_LOGS").select("*").execute()
            audit_data = audit_query.data if audit_query.data else []
            import pandas as pd

            audit_df = pd.DataFrame(audit_data)
        except Exception as e:
            st.error(f"Error initializing report data streams: {e}")
            fleet_df = pd.DataFrame()
            audit_df = pd.DataFrame()

        if not fleet_df.empty:
            # --- SANITIZE NUMERICAL COLUMNS FOR REPORTING ACCURACY ---
            if 'GEN_KVA' in fleet_df.columns:
                fleet_df['GEN_KVA'] = pd.to_numeric(fleet_df['GEN_KVA'], errors='coerce').fillna(0).astype(int)
            if 'RUN_HRS' in fleet_df.columns:
                fleet_df['RUN_HRS'] = pd.to_numeric(fleet_df['RUN_HRS'], errors='coerce').fillna(0).astype(int)

            # --- 2. GLOBAL EXECUTIVE FLEET METRICS ---
            st.markdown("##### 📊 Global Fleet Summary Metrics")

            total_units = len(fleet_df)
            total_capacity = fleet_df['GEN_KVA'].sum() if 'GEN_KVA' in fleet_df.columns else 0
            avg_hours = fleet_df['RUN_HRS'].mean() if 'RUN_HRS' in fleet_df.columns else 0
            total_incidents = len(audit_df) if not audit_df.empty else 0

            rep_kpi1, rep_kpi2, rep_kpi3, rep_kpi4 = st.columns(4)
            with rep_kpi1:
                st.metric(label="📟 ACTIVE DEPLOYED FLEET", value=f"{total_units:,} Units", border=True)
            with rep_kpi2:
                st.metric(label="⚡ TOTAL CAPACITY METRIC", value=f"{total_capacity:,} KVA", border=True)
            with rep_kpi3:
                st.metric(label="⏳ AVG RUNNING TIMELINE", value=f"{int(avg_hours):,} Hrs", border=True)
            with rep_kpi4:
                st.metric(label="📜 AUDITED SYSTEM EVENTS", value=f"{total_incidents:,} Actions", border=True)

            st.markdown("---")

            # --- 3. FILTER ENGINE WORKBENCH ---
            st.markdown("##### 🛠️ Report Generation Filter Engine")
            filter_col1, filter_col2, filter_col3 = st.columns(3)

            with filter_col1:
                user_options = ["ALL GROUPS"] + sorted(
                    fleet_df['USER'].dropna().unique().tolist()) if 'USER' in fleet_df.columns else ["ALL GROUPS"]
                selected_report_user = st.selectbox("Select Operational Group (USER):", options=user_options,
                                                    key="rep_filter_user")

            with filter_col2:
                loc_options = ["ALL LOCATIONS"] + sorted(
                    fleet_df['FIELD'].dropna().unique().tolist()) if 'FIELD' in fleet_df.columns else ["ALL LOCATIONS"]
                selected_report_field = st.selectbox("Select Field Location Assignment:", options=loc_options,
                                                     key="rep_filter_field")

            with filter_col3:
                type_options = ["ALL EQUIPMENT TYPES"] + sorted(
                    fleet_df['TYPE'].dropna().unique().tolist()) if 'TYPE' in fleet_df.columns else [
                    "ALL EQUIPMENT TYPES"]
                selected_report_type = st.selectbox("Select Generator Equipment Type:", options=type_options,
                                                    key="rep_filter_type")

            # Apply Processing Filters
            filtered_report_df = fleet_df.copy()
            if selected_report_user != "ALL GROUPS":
                filtered_report_df = filtered_report_df[filtered_report_df['USER'] == selected_report_user]
            if selected_report_field != "ALL LOCATIONS":
                filtered_report_df = filtered_report_df[filtered_report_df['FIELD'] == selected_report_field]
            if selected_report_type != "ALL EQUIPMENT TYPES":
                filtered_report_df = filtered_report_df[filtered_report_df['TYPE'] == selected_report_type]

            # --- DYNAMIC OPERATIONAL PROFILE BREAKDOWNS ---
            st.markdown("##### 📈 Granular Operational Analysis Profiles")
            an_col1, an_col2, an_col3 = st.columns(3)

            with an_col1:
                st.markdown("**🛠️ Fleet Count by Reason**")
                if 'REASON' in filtered_report_df.columns:
                    reason_counts = filtered_report_df['REASON'].fillna(
                        'Routine/Unspecified Operations').value_counts().to_frame().rename(
                        columns={'count': 'Unit Count'})
                    st.dataframe(reason_counts, use_container_width=True)
                else:
                    st.caption("No explicit REASON logs found.")

            with an_col2:
                st.markdown("**🚚 Logistical Transfer Status**")
                if 'TO_LOCATION' in filtered_report_df.columns:
                    transfer_counts = filtered_report_df['TO_LOCATION'].fillna(
                        'Stationary / Located').value_counts().to_frame().rename(columns={'count': 'Unit Count'})
                    st.dataframe(transfer_counts, use_container_width=True)
                else:
                    st.caption("No movement tracking column found.")

            with an_col3:
                st.markdown("**🎯 Operational Allocation Purpose**")
                purpose_col = 'PURPOSE' if 'PURPOSE' in filtered_report_df.columns else (
                    'REMARKS' if 'REMARKS' in filtered_report_df.columns else None)
                if purpose_col:
                    purpose_counts = filtered_report_df[purpose_col].fillna(
                        'General Deployment Support').value_counts().to_frame().rename(columns={'count': 'Unit Count'})
                    st.dataframe(purpose_counts, use_container_width=True)
                else:
                    st.caption("No structural PURPOSE data fields present.")

            st.markdown("---")

            # --- 4. DATA EXPORT STUDIO ---
            st.markdown("##### 💾 Professional Document Export Studio")

            export_mode = st.radio(
                "Choose Target Presentation Document Format:",
                options=["Excel Spreadsheet (.xlsx)", "Print-Ready Document (.pdf)"],
                horizontal=True,
                key="document_export_mode_selector"
            )

            import io
            import datetime


            # Safe String Transformer to handle character sets safely
            def safe_str(val):
                if val is None or pd.isna(val) or str(val).strip().upper() == "NONE":
                    return "-"
                return str(val).strip().encode('ascii', 'ignore').decode('ascii')


            if export_mode == "Excel Spreadsheet (.xlsx)":
                import openpyxl
                from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
                from openpyxl.utils import get_column_letter

                excel_buffer = io.BytesIO()
                wb = openpyxl.Workbook()

                # Tab 1: Executive Summary
                ws_dash = wb.active
                ws_dash.title = "Executive Summary"
                ws_dash.views.sheetView[0].showGridLines = True

                navy_dark, accent_blue, zebra_tint, white, b_gray = "1B365D", "E8EEF5", "F7F9FB", "FFFFFF", "D3D3D3"

                ws_dash.merge_cells("A1:E2")
                t_cell = ws_dash["A1"]
                t_cell.value = "FIELD OPERATIONS EXECUTIVES SUMMARY PROFILE"
                t_cell.font = Font(name="Calibri", size=14, bold=True, color=white)
                t_cell.fill = PatternFill(start_color=navy_dark, end_color=navy_dark, fill_type="solid")
                t_cell.alignment = Alignment(horizontal="center", vertical="center")

                ws_dash["A3"] = f"Generated: {datetime.date.today().isoformat()} | Group Scope: {selected_report_user}"
                ws_dash["A3"].font = Font(italic=True, size=10)

                kpis = [
                    ("Active Units", f"{len(filtered_report_df)} Pcs", "A"),
                    ("Total KVA Rating", f"{filtered_report_df['GEN_KVA'].sum():,} KVA", "B"),
                    ("Avg Hours Registered",
                     f"{int(filtered_report_df['RUN_HRS'].mean() if len(filtered_report_df) > 0 else 0):,} Hrs", "C")
                ]

                for title, val, col in kpis:
                    ws_dash[f"{col}5"] = title
                    ws_dash[f"{col}5"].font = Font(bold=True, size=11, color=navy_dark)
                    ws_dash[f"{col}5"].fill = PatternFill(start_color=accent_blue, fill_type="solid")
                    ws_dash[f"{col}5"].alignment = Alignment(horizontal="center")

                    ws_dash[f"{col}6"] = val
                    ws_dash[f"{col}6"].font = Font(bold=True, size=13)
                    ws_dash[f"{col}6"].alignment = Alignment(horizontal="center")

                    thin = Side(border_style="thin", color=b_gray)
                    ws_dash[f"{col}5"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    ws_dash[f"{col}6"].border = Border(top=thin, left=thin, right=thin, bottom=thin)

                # Tab 2: Master Live Dataset
                ws_data = wb.create_sheet(title="Master Live Dataset")
                ws_data.views.sheetView[0].showGridLines = True

                headers = list(filtered_report_df.columns)
                for c_idx, h_text in enumerate(headers, 1):
                    cell = ws_data.cell(row=1, column=c_idx, value=str(h_text).upper())
                    cell.font = Font(name="Calibri", size=11, bold=True, color=white)
                    cell.fill = PatternFill(start_color=navy_dark, end_color=navy_dark, fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                t_border = Border(left=Side(style='thin', color=b_gray), right=Side(style='thin', color=b_gray),
                                  top=Side(style='thin', color=b_gray), bottom=Side(style='thin', color=b_gray))

                for r_idx, r_values in enumerate(filtered_report_df.values, 2):
                    row_fill = PatternFill(start_color=zebra_tint,
                                           fill_type="solid") if r_idx % 2 == 0 else PatternFill(fill_type=None)
                    for c_idx, val in enumerate(r_values, 1):
                        write_val = safe_str(val) if isinstance(val, str) else val
                        cell = ws_data.cell(row=r_idx, column=c_idx, value=write_val)
                        cell.fill = row_fill
                        cell.border = t_border
                        if isinstance(val, (int, float)):
                            cell.number_format = '#,##0'
                            cell.alignment = Alignment(horizontal="right")

                for sheet in [ws_dash, ws_data]:
                    for col in sheet.columns:
                        m_len = max(len(str(cell.value or '')) for cell in col)
                        col_letter = get_column_letter(col[0].column)
                        sheet.column_dimensions[col_letter].width = max(m_len + 4, 12)

                wb.save(excel_buffer)
                st.download_button(
                    label="🟢 DOWNLOAD DESIGNED EXCEL WORKBOOK (.XLSX)",
                    data=excel_buffer.getvalue(),
                    file_name=f"Executive_Fleet_Report_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )


            elif export_mode == "Print-Ready Document (.pdf)":

                def generate_fleet_pdf(dataframe, user_scope):

                    from fpdf import FPDF

                    class CustomFleetPDF(FPDF):

                        def header(self):
                            self.set_fill_color(27, 54, 93)

                            self.rect(10, 10, 277, 24, "F")

                            self.set_text_color(255, 255, 255)

                            self.set_font("Helvetica", "B", 13)

                            self.set_y(14)

                            self.cell(0, 8, "   FIELD OPERATIONS FLEET SUMMARY REPORT", ln=True)

                            self.set_font("Helvetica", "I", 9)

                            self.cell(0, 4,
                                      f"   Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')} | Scope: {user_scope}",
                                      ln=True)

                            self.set_y(40)

                        def footer(self):
                            self.set_y(-15)

                            self.set_font("Helvetica", "I", 8)

                            self.set_text_color(128, 128, 128)

                            self.cell(0, 10, f"Page {self.page_no()}/{{nb}}", align="C")

                    pdf = CustomFleetPDF(orientation="L", unit="mm", format="A4")

                    pdf.alias_nb_pages()

                    pdf.add_page()

                    # Table headers start directly here without metric cards

                    pdf.set_xy(10, 42)

                    pdf.set_fill_color(27, 54, 93)

                    pdf.set_text_color(255, 255, 255)

                    pdf.set_font("Helvetica", "B", 8)

                    col_widths = [16, 23, 19, 14, 15, 28, 30, 38, 50, 44]

                    headers = ["G-CODE", "TYPE", "MODEL", "KVA", "HOURS", "STATUS/USER", "FIELD SITE", "TRANSFER STAT",
                               "LOGGED REASON", "PURPOSE TARGET"]

                    for w, h in zip(col_widths, headers):
                        pdf.cell(w, 8, h, border=1, align="C", fill=True)

                    pdf.ln()

                    pdf.set_font("Helvetica", "", 7.5)

                    pdf.set_text_color(50, 50, 50)

                    fill_toggle = False

                    for index, row in dataframe.iterrows():

                        if pdf.get_y() > 180:

                            pdf.add_page()

                            pdf.set_xy(10, 42)

                            pdf.set_fill_color(27, 54, 93)

                            pdf.set_text_color(255, 255, 255)

                            pdf.set_font("Helvetica", "B", 8)

                            for w, h in zip(col_widths, headers):
                                pdf.cell(w, 8, h, border=1, align="C", fill=True)

                            pdf.ln()

                            pdf.set_font("Helvetica", "", 7.5)

                            pdf.set_text_color(50, 50, 50)

                        pdf.set_fill_color(247, 249, 251) if fill_toggle else pdf.set_fill_color(255, 255, 255)

                        pdf.cell(col_widths[0], 7, safe_str(row.get('G-CODE')), border=1, fill=True, align="C")

                        pdf.cell(col_widths[1], 7, safe_str(row.get('TYPE'))[:14], border=1, fill=True)

                        pdf.cell(col_widths[2], 7, safe_str(row.get('MODEL'))[:12], border=1, fill=True)

                        pdf.cell(col_widths[3], 7, f"{int(row.get('GEN_KVA', 0)):,}", border=1, fill=True, align="R")

                        pdf.cell(col_widths[4], 7, f"{int(row.get('RUN_HRS', 0)):,}", border=1, fill=True, align="R")

                        pdf.cell(col_widths[5], 7, safe_str(row.get('USER'))[:18], border=1, fill=True)

                        pdf.cell(col_widths[6], 7, safe_str(row.get('FIELD'))[:18], border=1, fill=True)

                        pdf.cell(col_widths[7], 7, safe_str(row.get('TO_LOCATION'))[:22], border=1, fill=True)

                        pdf.cell(col_widths[8], 7, safe_str(row.get('REASON'))[:32], border=1, fill=True)

                        p_val = row.get('PURPOSE', row.get('REMARKS', '-'))

                        pdf.cell(col_widths[9], 7, safe_str(p_val)[:26], border=1, fill=True)

                        pdf.ln()

                        fill_toggle = not fill_toggle

                    # 💡 FIX: Safely grab raw string bytes directly from the compilation layer

                    return pdf.output(dest='S').encode('latin-1')


                if filtered_report_df.empty:

                    st.warning("⚠️ No database items match your current filter parameters to render onto a PDF canvas.")

                else:

                    pdf_data = generate_fleet_pdf(filtered_report_df, selected_report_user)

                    st.download_button(

                        label="🔴 DOWNLOAD PRINT-READY REPORT DOCUMENT (.PDF)",

                        data=pdf_data,

                        file_name=f"Executive_Fleet_Report_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.pdf",

                        mime="application/pdf",

                        use_container_width=True

                    )

            st.markdown("---")

            # --- 5. RENDER FULL MASTER REPORT DATAFRAME VIEWPORT ---
            st.markdown(
                f"##### 📋 Generated Master Ledger Viewport ({len(filtered_report_df)} Units Matching Selections)")

            if not filtered_report_df.empty:
                report_display_df = filtered_report_df.rename(columns={
                    'G-CODE': 'Asset ID (G-CODE)',
                    'TYPE': 'Equipment Type',
                    'MODEL': 'Model Designation',
                    'GEN_KVA': 'Capacity (KVA)',
                    'RUN_HRS': 'Accumulated Run Hours',
                    'USER': 'Operational Status/User',
                    'FIELD': 'Assigned Field Site',
                    'TO_LOCATION': 'Transfer Status (Target)',
                    'MOVE_DATE': 'Last Movement Date',
                    'REASON': 'Logged Reason',
                    'PURPOSE': 'Allocation Purpose'
                })

                target_report_cols = [
                    'Asset ID (G-CODE)', 'Equipment Type', 'Model Designation', 'Capacity (KVA)',
                    'Accumulated Run Hours', 'Operational Status/User', 'Assigned Field Site',
                    'Transfer Status (Target)', 'Logged Reason', 'Allocation Purpose'
                ]

                available_report_cols = [c for c in target_report_cols if c in report_display_df.columns]
                styled_report_grid = report_display_df[available_report_cols].style.apply(style_zebra_rows, axis=None)

                st.dataframe(
                    styled_report_grid,
                    use_container_width=True,
                    hide_index=True,
                    column_config={
                        "Capacity (KVA)": st.column_config.NumberColumn(format="%d"),
                        "Accumulated Run Hours": st.column_config.NumberColumn(format="%d")
                    }
                )
            else:
                st.warning(
                    "⚠️ No active asset matches found inside database sheets for the selected filter combinations.")

        else:
            st.info("No equipment inventory assets found inside database registries to evaluate.")
